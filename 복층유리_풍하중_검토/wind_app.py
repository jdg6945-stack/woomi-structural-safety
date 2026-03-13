# -*- coding: utf-8 -*-
"""단위세대 복층유리 두께 산정 프로그램 — Streamlit 변환판 (Wind Checker v4)"""
import os
import sys
import io
import re
import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# calculator.py를 같은 폴더에서 import
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from calculator import GlassCalculator

# ─── PDF 생성용 (reportlab) ──────────────────────
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors


RECOMMEND_PRICE_ROWS = [
    {"rank": 1, "spec": "5mm(AN) + Air + 5mm(AN)", "rec1_default": 21000},
    {"rank": 2, "spec": "5mm(HS) + Air + 5mm(AN)", "rec1_default": 26000},
    {"rank": 3, "spec": "5mm(HS) + Air + 5mm(HS)", "rec1_default": 30000},
    {"rank": 4, "spec": "6mm(AN) + Air + 6mm(AN)", "rec1_default": 30000},
    {"rank": 5, "spec": "6mm(HS) + Air + 6mm(AN)", "rec1_default": 35000},
    {"rank": 6, "spec": "6mm(HS) + Air + 6mm(HS)", "rec1_default": 40000},
    {"rank": 7, "spec": "8mm(AN) + Air + 8mm(AN)", "rec1_default": None},
    {"rank": 8, "spec": "8mm(HS) + Air + 8mm(AN)", "rec1_default": None},
    {"rank": 9, "spec": "8mm(HS) + Air + 8mm(HS)", "rec1_default": None},
    {"rank": 10, "spec": "5mm(FT) + Air + 5mm(AN)", "rec1_default": None},
    {"rank": 11, "spec": "6mm(FT) + Air + 6mm(AN)", "rec1_default": None},
    {"rank": 12, "spec": "8mm(FT) + Air + 8mm(AN)", "rec1_default": None},
    {"rank": 13, "spec": "5mm(FT) + Air + 5mm(HS)", "rec1_default": None},
    {"rank": 14, "spec": "6mm(FT) + Air + 6mm(HS)", "rec1_default": None},
    {"rank": 15, "spec": "8mm(FT) + Air + 8mm(HS)", "rec1_default": None},
    {"rank": 16, "spec": "5mm(FT) + Air + 5mm(FT)", "rec1_default": None},
    {"rank": 17, "spec": "6mm(FT) + Air + 6mm(FT)", "rec1_default": None},
    {"rank": 18, "spec": "8mm(FT) + Air + 8mm(FT)", "rec1_default": None},
]


def _default_rec1_prices() -> dict:
    return {row["spec"]: row.get("rec1_default") for row in RECOMMEND_PRICE_ROWS}


def _default_rec2_prices() -> dict:
    return {row["spec"]: None for row in RECOMMEND_PRICE_ROWS}


# ==========================================
# PDF 유틸 함수 (호출 전에 정의)
# ==========================================
def _register_pdf_fonts():
    """PDF용 한글 폰트 등록"""
    _dir = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(_dir, "NanumGothic.ttf"),
        os.path.join(_dir, "malgunbd.ttf"),
        r"C:\Windows\Fonts\malgun.ttf",
        r"C:\Windows\Fonts\NanumGothic.ttf",
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf",
    ]
    reg_path, bold_path = None, None
    for p in candidates:
        if not os.path.exists(p):
            continue
        n = os.path.basename(p).lower()
        if n in ("malgun.ttf", "nanumgothic.ttf") and reg_path is None:
            reg_path = p
        if n in ("malgunbd.ttf", "nanumgothicbold.ttf") and bold_path is None:
            bold_path = p
    if not reg_path:
        reg_path = bold_path
    if not bold_path:
        bold_path = reg_path
    if not reg_path:
        return None, None
    try:
        pdfmetrics.registerFont(TTFont("KR-Regular", reg_path))
    except Exception:
        pass
    try:
        pdfmetrics.registerFont(TTFont("KR-Bold", bold_path))
    except Exception:
        pdfmetrics.registerFont(TTFont("KR-Bold", reg_path))
    return "KR-Regular", "KR-Bold"


def _generate_pdf(content: str) -> bytes:
    """detail 텍스트를 PDF 바이트로 변환"""
    BASE_FONT, BOLD_FONT = _register_pdf_fonts()
    if not BASE_FONT:
        return None

    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    c.setFont(BASE_FONT, 9)
    c.drawRightString(width - 40, height - 40, "Wind Checker-Version4_우미건설 품질기술팀")
    c.setFont(BOLD_FONT, 12)
    c.drawString(40, height - 54, "Wind Checker (단위세대 복층유리 두께 산정 프로그램)")
    c.setFont(BASE_FONT, 10)

    date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    x_margin = 40
    y = height - 80
    line_height = 14
    group_line_re = re.compile(r"^\d+\s*층\s*~\s*\d+\s*층\s*:")

    def draw_line(line: str):
        nonlocal y
        if y < 60:
            c.showPage()
            c.setFont(BASE_FONT, 9)
            c.drawRightString(width - 40, height - 40, "Wind Checker-Version4_우미건설 품질기술팀")
            c.setFont(BASE_FONT, 10)
            y = height - 80

        line = line.replace("!!BLUEBOLD!!", "")

        if line.startswith("▶ 검토 결과(층별 추천_그룹화)"):
            c.setFont(BOLD_FONT, 10)
            c.setFillColor(colors.blue)
            c.drawString(x_margin, y, line)
            c.setFillColor(colors.black)
            c.setFont(BASE_FONT, 10)
            y -= line_height
            return
        if group_line_re.match(line):
            c.setFont(BOLD_FONT, 10)
            c.setFillColor(colors.blue)
            c.drawString(x_margin, y, line)
            c.setFillColor(colors.black)
            c.setFont(BASE_FONT, 10)
            y -= line_height
            return
        if line.startswith("# "):
            c.setFont(BOLD_FONT, 10)
            c.drawString(x_margin, y, line)
            c.setFont(BASE_FONT, 10)
            y -= line_height
            return
        if line.startswith("**") and line.endswith("**") and len(line) >= 4:
            c.setFont(BOLD_FONT, 10)
            c.drawString(x_margin, y, line[2:-2])
            c.setFont(BASE_FONT, 10)
            y -= line_height
            return
        c.drawString(x_margin, y, line)
        y -= line_height

    for raw in content.strip().splitlines():
        draw_line(raw.rstrip())

    c.setFont(BASE_FONT, 9)
    c.drawRightString(width - 40, 30, date_str)
    c.save()
    return buf.getvalue()
def _load_manual_pdf() -> tuple[bytes | None, str]:
    """사용 메뉴얼 PDF를 읽어 바이트와 다운로드 파일명을 반환한다."""
    manual_name = "Wind Checker_rev4 사용 매뉴얼_웹기반(2603).pdf"
    manual_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), manual_name)
    if not os.path.exists(manual_path):
        return None, manual_name
    with open(manual_path, "rb") as manual_file:
        return manual_file.read(), manual_name


def _extract_floor_recommendations(brief: str, num_floors: int) -> dict:
    """brief 문자열에서 층별 추천 스펙을 추출한다."""
    floor_specs = {f: "-" for f in range(1, int(num_floors) + 1)}
    in_result = False
    range_re = re.compile(r"^(\d+)\s*층\s*~\s*(\d+)\s*층\s*:\s*(.+)$")
    single_re = re.compile(r"^(\d+)\s*층\s*:\s*(.+)$")

    for raw in brief.splitlines():
        line = raw.strip()
        if not line:
            continue
        if "▶ 검토 결과" in line:
            in_result = True
            continue
        if in_result and line.startswith("▶ "):
            break
        if not in_result:
            continue

        m_range = range_re.match(line)
        if m_range:
            top_f = int(m_range.group(1))
            low_f = int(m_range.group(2))
            spec = m_range.group(3).strip()
            for f in range(min(low_f, top_f), max(low_f, top_f) + 1):
                if 1 <= f <= int(num_floors):
                    floor_specs[f] = spec
            continue

        m_single = single_re.match(line)
        if m_single:
            f = int(m_single.group(1))
            spec = m_single.group(2).strip()
            if 1 <= f <= int(num_floors):
                floor_specs[f] = spec

    return floor_specs


def _get_spec_fill(spec: str) -> PatternFill:
    """조합 문자열 기준으로 고정 배경색을 반환한다."""
    key = (spec or "").replace(" ", "").upper()

    fixed_map = {
        "5MM(AN)+AIR+5MM(AN)": "FFFFFF",
        "5MM(HS)+AIR+5MM(AN)": "C9DAF8",
        "6MM(AN)+AIR+6MM(AN)": "FCE5CD",
        "5MM(HS)+AIR+5MM(HS)": "9FC5E8",
        "6MM(HS)+AIR+6MM(AN)": "F4B183",
        "6MM(HS)+AIR+6MM(HS)": "C55A11",
        "5MM(FT)+AIR+5MM(FT)": "1F4E78",
        "6MM(FT)+AIR+6MM(FT)": "843C0C",
    }

    if key in fixed_map:
        return PatternFill("solid", fgColor=fixed_map[key])

    # 정의되지 않은 조합도 항상 같은 색이 되도록 고정 팔레트에서 결정
    fallback_palette = [
        "FFF2CC", "D9EAD3", "D0E0E3", "EAD1DC",
        "F9CB9C", "CFE2F3", "E6B8AF", "D9D2E9",
    ]
    idx = sum(ord(ch) for ch in key) % len(fallback_palette) if key else 0
    return PatternFill("solid", fgColor=fallback_palette[idx])


def _get_active_combo_price_map(data: dict) -> dict:
    """현재 추천 옵션 기준 조합 단가표를 반환한다."""
    unit_prices = data.get("단가표", {}) if isinstance(data, dict) else {}
    if not isinstance(unit_prices, dict):
        return {}

    rec_opt = str(data.get("유리 Spec 추천 옵션", "추천1")).strip()
    if rec_opt == "추천1":
        table = unit_prices.get("rec1_combo_prices", {})
    elif rec_opt == "추천2":
        table = unit_prices.get("rec2_combo_prices", {})
    else:
        # 추천3은 구조 중심이므로 단가표는 추천2 -> 추천1 순으로 fallback
        table = unit_prices.get("rec2_combo_prices", {}) or unit_prices.get("rec1_combo_prices", {})

    if not isinstance(table, dict) or not table:
        table = unit_prices.get("combo_prices", {})

    return table if isinstance(table, dict) else {}


def _build_area_cost_summary(calc_results: list) -> dict:
    """검토 결과에서 조합별 면적/단가/예상금액 집계를 생성한다."""
    if not calc_results:
        return {
            "rows": [],
            "total_area": 0.0,
            "total_amount": 0.0,
            "missing_specs": [],
            "price_map": {},
        }

    first = calc_results[0]
    data = first.get("data", {})
    try:
        num_floors = int(float(data.get("건물층수", 0) or 0))
    except Exception:
        num_floors = 0

    spec_rank = {row["spec"]: int(row["rank"]) for row in RECOMMEND_PRICE_ROWS}
    price_map = _get_active_combo_price_map(data)
    area_by_spec = {}

    for item in calc_results:
        d = item.get("data", {})
        brief = item.get("outs", {}).get("brief", "")
        floor_specs = _extract_floor_recommendations(brief, num_floors)
        try:
            area = float(d.get("유리 폭 (m)", 0)) * float(d.get("유리 높이 (m)", 0))
        except Exception:
            area = 0.0
        if area <= 0:
            continue

        for _, spec in floor_specs.items():
            s = (spec or "").strip()
            if not s or s in ("-", "FAIL"):
                continue
            area_by_spec[s] = area_by_spec.get(s, 0.0) + area

    ordered_specs = sorted(area_by_spec.keys(), key=lambda s: (spec_rank.get(s, 999), s))
    rows = []
    total_area = 0.0
    total_amount = 0.0
    missing_specs = []

    for spec in ordered_specs:
        area = float(area_by_spec.get(spec, 0.0))
        total_area += area

        raw_price = price_map.get(spec, None)
        price = None
        if raw_price is not None and str(raw_price).strip() != "":
            try:
                p = float(raw_price)
                if p > 0:
                    price = p
            except Exception:
                price = None

        amount = None
        if price is None:
            missing_specs.append(spec)
        else:
            amount = area * price
            total_amount += amount

        rows.append({
            "spec": spec,
            "area": area,
            "price": price,
            "amount": amount,
        })

    return {
        "rows": rows,
        "total_area": total_area,
        "total_amount": total_amount,
        "missing_specs": missing_specs,
        "price_map": price_map,
    }


def _build_area_cost_summary_for_item(item: dict, all_calc_results: list) -> dict:
    """단일 유리(calc result) 항목에 대한 면적/단가/예상금액 집계를 생성한다."""
    if not item:
        return {
            "rows": [],
            "total_area": 0.0,
            "total_amount": 0.0,
            "missing_specs": [],
        }

    first_result = all_calc_results[0] if all_calc_results else item
    data = first_result.get("data", {})
    try:
        num_floors = int(float(data.get("건물층수", 0) or 0))
    except Exception:
        num_floors = 0

    spec_rank = {row["spec"]: int(row["rank"]) for row in RECOMMEND_PRICE_ROWS}
    price_map = _get_active_combo_price_map(first_result.get("data", {}))

    d = item.get("data", {})
    brief = item.get("outs", {}).get("brief", "")
    floor_specs = _extract_floor_recommendations(brief, num_floors)
    try:
        area = float(d.get("유리 폭 (m)", 0)) * float(d.get("유리 높이 (m)", 0))
    except Exception:
        area = 0.0

    if area <= 0:
        return {
            "rows": [],
            "total_area": 0.0,
            "total_amount": 0.0,
            "missing_specs": [],
        }

    # 이 유리에서 사용되는 조합들 수집
    unique_specs = set()
    for _, spec in floor_specs.items():
        s = (spec or "").strip()
        if s and s not in ("-", "FAIL"):
            unique_specs.add(s)

    ordered_specs = sorted(unique_specs, key=lambda s: (spec_rank.get(s, 999), s))
    rows = []
    total_area = 0.0
    total_amount = 0.0
    missing_specs = []

    for spec in ordered_specs:
        # 각 조합이 이 유리에서 차지하는 비율 계산 (층별로 균등 분배)
        num_layers_with_spec = sum(1 for _, s in floor_specs.items() if (s or "").strip() == spec)
        if num_layers_with_spec == 0:
            continue
        spec_area = area * num_layers_with_spec / max(1, len([s for s in floor_specs.values() if (s or "").strip() not in ("-", "FAIL", "")]))

        raw_price = price_map.get(spec, None)
        price = None
        if raw_price is not None and str(raw_price).strip() != "":
            try:
                p = float(raw_price)
                if p > 0:
                    price = p
            except Exception:
                price = None

        amount = None
        if price is None:
            missing_specs.append(spec)
        else:
            amount = spec_area * price
            total_amount += amount

        total_area += spec_area

        rows.append({
            "spec": spec,
            "area": spec_area,
            "price": price,
            "amount": amount,
        })

    return {
        "rows": rows,
        "total_area": total_area,
        "total_amount": total_amount,
        "missing_specs": missing_specs,
    }


def _generate_excel(calc_results: list) -> bytes:
    """요청 폼 기준: 유리별 열 확장 + 층별 스펙 변경시 구분색 적용"""
    if not calc_results:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "검토결과"

    title_font = Font(size=15, bold=True)
    head_font = Font(size=11, bold=True)
    base_font = Font(size=10)
    blue_font = Font(size=10, color="1D4ED8", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_head = PatternFill("solid", fgColor="EAEAEA")

    first = calc_results[0]
    data = first.get("data", {})
    num_floors = int(float(data.get("건물층수", 0) or 0))
    v0 = data.get("기본풍속 (V0) [m/sec]", "-")
    exposure = data.get("노풍도 (A, B, C, D)", "-")
    rec_opt = data.get("유리 Spec 추천 옵션", "-")

    # 컬럼 설정: A=구분, B..=유리별
    ws.column_dimensions["A"].width = 16
    for idx in range(len(calc_results)):
        col = chr(ord("B") + idx)
        ws.column_dimensions[col].width = 24

    ws["A1"] = "XX현장 유리두께 정리"
    ws["A1"].font = title_font
    ws["A2"] = "건물층수:"
    ws["B2"] = f"{num_floors}층"
    ws["A3"] = "기본풍속:"
    ws["B3"] = f"{v0} m/s"
    ws["A4"] = "노풍도:"
    ws["B4"] = f"{exposure}"
    ws["A5"] = "Spec추천"
    ws["B5"] = f"{rec_opt}"

    for c in ["A2", "A3", "A4", "A5"]:
        ws[c].font = head_font
    for c in ["B2", "B3", "B4", "B5"]:
        ws[c].font = base_font


    ws.row_dimensions[6].height = 8

    # 표 헤더
    head_row = 8
    ws[f"A{head_row}"] = "구분"
    ws[f"A{head_row}"].font = head_font
    ws[f"A{head_row}"].alignment = center
    ws[f"A{head_row}"].fill = fill_head
    ws[f"A{head_row}"].border = border

    for idx, item in enumerate(calc_results):
        col = chr(ord("B") + idx)
        ws[f"{col}{head_row}"] = item.get("label", f"유리{idx + 1}")
        ws[f"{col}{head_row}"].font = head_font
        ws[f"{col}{head_row}"].alignment = center
        ws[f"{col}{head_row}"].fill = fill_head
        ws[f"{col}{head_row}"].border = border

    row_size = head_row + 1
    row_area = head_row + 2
    row_pos = head_row + 3
    row_floor_header = head_row + 4

    ws[f"A{row_size}"] = "유리크기(mm)"
    ws[f"A{row_area}"] = "유리면적(m2)"
    ws[f"A{row_pos}"] = "검토위치"
    ws[f"A{row_floor_header}"] = "층"
    for r in [row_size, row_area, row_pos, row_floor_header]:
        ws[f"A{r}"].font = head_font if r == row_floor_header else base_font
        ws[f"A{r}"].alignment = center
        ws[f"A{r}"].border = border
        if r == row_floor_header:
            ws[f"A{r}"].fill = fill_head

    floor_specs_by_glass = []
    for idx, item in enumerate(calc_results):
        col = chr(ord("B") + idx)
        d = item.get("data", {})
        brief = item.get("outs", {}).get("brief", "")
        floor_specs = _extract_floor_recommendations(brief, num_floors)
        floor_specs_by_glass.append(floor_specs)

        try:
            gw = float(d.get("유리 폭 (m)", 0))
            gh = float(d.get("유리 높이 (m)", 0))
        except Exception:
            gw, gh = 0.0, 0.0

        ws[f"{col}{row_size}"] = f"{int(round(gw * 1000))} x {int(round(gh * 1000))}"
        ws[f"{col}{row_area}"] = f"{gw * gh:.2f}"
        ws[f"{col}{row_pos}"] = d.get("검토위치 (중앙부, 모서리부)", "-")
        ws[f"{col}{row_floor_header}"] = "추천 Spec"

        for r in [row_size, row_area, row_pos, row_floor_header]:
            ws[f"{col}{r}"].alignment = center if r != row_floor_header else center
            ws[f"{col}{r}"].font = head_font if r == row_floor_header else base_font
            ws[f"{col}{r}"].border = border
            if r == row_floor_header:
                ws[f"{col}{r}"].fill = fill_head

    # 층별 결과 채우기: 같은 조합은 열과 무관하게 동일 색상 적용
    start_row = row_floor_header + 1
    for r_idx, f in enumerate(range(num_floors, 0, -1)):
        row = start_row + r_idx
        ws[f"A{row}"] = f"{f}층"
        ws[f"A{row}"].alignment = center
        ws[f"A{row}"].font = base_font
        ws[f"A{row}"].border = border

        for idx, floor_specs in enumerate(floor_specs_by_glass):
            col = chr(ord("B") + idx)
            spec = floor_specs.get(f, "-")
            ws[f"{col}{row}"] = spec
            ws[f"{col}{row}"].alignment = left
            ws[f"{col}{row}"].font = blue_font
            ws[f"{col}{row}"].border = border
            ws[f"{col}{row}"].fill = _get_spec_fill(spec)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ==========================================
# 0. 페이지 설정 & 스타일
# ==========================================
try:
    st.set_page_config(layout="wide", page_title="Wind Checker v4 - 단위세대 복층유리 두께 산정 프로그램", page_icon="🌬️")
except Exception:
    pass

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700;900&display=swap');
.stApp { font-family: 'Noto Sans KR', sans-serif !important; background-color: #edf2f7; }
[data-testid="stAppViewContainer"] {
    background-color: #edf2f7 !important;
}
[data-testid="stAppViewContainer"] .main {
    background-color: #edf2f7 !important;
}
[data-testid="stAppViewContainer"] .block-container {
    background-color: #f4f7fb;
    border-radius: 10px;
    padding-top: 1.8rem;
    padding-bottom: 2rem;
}
.stMarkdown, .stMarkdown p, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3,
[data-testid="stSidebar"], [data-testid="stHeader"],
.stTextInput input, .stNumberInput input, .stSelectbox, .stMultiSelect {
    font-family: 'Noto Sans KR', sans-serif !important;
}
[data-testid="stSidebar"] { background-color: #e3e9f1; border-right: 1px solid #c6d1e0; }
[data-testid="stSidebar"] .stMarkdown label,
[data-testid="stSidebar"] .stNumberInput label,
[data-testid="stSidebar"] .stSelectbox label {
    color: #1f2937 !important;
    font-size: 13px !important;
    font-weight: 600 !important;
}
[data-testid="stSidebar"] hr {
    border-top: 1px solid #cfd8e3 !important;
}
/* 사이드바 입력 필드 그룹 테두리 */
.sidebar-group {
    border: 1.5px solid #aebbd1;
    border-radius: 10px;
    padding: 14px;
    margin-bottom: 14px;
    background: #ffffff;
}
.sidebar-group-title {
    font-size: 14px; font-weight: 700; color: #1f2937; margin-bottom: 10px;
}
.sidebar-section-title {
    font-size: 24px; font-weight: 800; color: #0f172a; margin: 10px 0 8px 0;
    border-bottom: 1px solid #cfd8e3; padding-bottom: 8px;
}
/* 사이드바 입력창 자체 테두리 */
[data-testid="stSidebar"] div[data-baseweb="input"] {
    border: 1.2px solid #aebbd1 !important;
    border-radius: 8px !important;
    background: #ffffff !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] div[data-baseweb="input"]:focus-within {
    border-color: #5b7db4 !important;
    box-shadow: 0 0 0 2px rgba(91, 125, 180, 0.12) !important;
}
[data-testid="stSidebar"] div[data-baseweb="select"] > div {
    border: 1.2px solid #aebbd1 !important;
    border-radius: 8px !important;
    min-height: 40px !important;
    background: #ffffff !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] div[data-baseweb="select"] > div:focus-within {
    border-color: #5b7db4 !important;
    box-shadow: 0 0 0 2px rgba(91, 125, 180, 0.12) !important;
}
/* number_input의 +/- 버튼 숨김 */
[data-testid="stSidebar"] [data-testid="stNumberInput"] button {
    display: none !important;
}
[data-testid="stSidebar"] [data-testid="stNumberInput"] [data-baseweb="input"] {
    width: 100% !important;
}
[data-testid="stSidebar"] [data-testid="stNumberInput"] input {
    text-align: left !important;
}
.stDownloadButton > button {
    background: #dbeafe !important;
    color: #1e3a8a !important;
    border: 1px solid #93c5fd !important;
    border-radius: 8px !important;
}
.stDownloadButton > button:hover {
    background: #bfdbfe !important;
    color: #1e3a8a !important;
    border-color: #60a5fa !important;
}
/* 대시보드 패널 */
.panel-box {
    background: #ffffff; border: 1.5px solid #e2e8f0; border-radius: 12px;
    padding: 20px 22px; height: 100%;
}
.panel-title {
    font-size: 1.05rem; font-weight: 800; color: #1e293b; margin-bottom: 14px;
    padding-bottom: 8px; border-bottom: 2px solid #2563eb;
}
.cond-row { display: flex; justify-content: space-between; padding: 5px 0; font-size: 0.93rem; border-bottom: 1px dashed #f1f5f9; }
.cond-label { color: #64748b; }
.cond-value { font-weight: 700; color: #1e293b; }
.result-header {
    color: #1e40af; font-weight: 900; font-size: 1.05rem; margin-top: 10px;
}
.result-group {
    color: #1e40af; font-weight: 700; font-size: 0.98rem; padding: 2px 0;
}
.result-normal {
    font-size: 0.95rem; line-height: 1.6; color: #334155;
}
.guide-box {
    background: #f1f5f9; border-left: 4px solid #2563eb; padding: 14px 18px;
    border-radius: 8px; margin-bottom: 18px; font-size: 15px; line-height: 1.7;
}
.detail-table {
    width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 8px;
}
.detail-table th {
    background: #e2e8f0; padding: 6px 8px; text-align: center; font-weight: 700;
    border: 1px solid #cbd5e1;
}
.detail-table td {
    padding: 5px 8px; text-align: center; border: 1px solid #e2e8f0;
}
.detail-table tr:nth-child(even) { background: #f8fafc; }
.combo-ok { color: #16a34a; font-weight: 700; }
.combo-ng { color: #dc2626; }
.combo-picked { background: #dbeafe !important; font-weight: 700; }
.floor-popup-head {
    background: #ffffff;
    border: 1.5px solid #cfd8e3;
    border-radius: 12px;
    padding: 10px 12px;
    box-shadow: 0 10px 26px rgba(15, 23, 42, 0.10);
    margin-bottom: 10px;
}
.floor-popup-title {
    font-size: 1rem;
    font-weight: 800;
    color: #1e293b;
    margin-bottom: 4px;
}
.floor-popup-sub {
    font-size: 12px;
    color: #64748b;
}
.summary-wrap {
    background: #ffffff;
    border: 1.5px solid #dbe4f0;
    border-radius: 12px;
    padding: 14px;
    margin-bottom: 12px;
}
.summary-grid {
    display: grid;
    grid-template-columns: repeat(3, minmax(0, 1fr));
    gap: 10px;
}
.summary-card {
    background: #f8fbff;
    border: 1px solid #d6e4f5;
    border-radius: 10px;
    padding: 10px 12px;
}
.summary-label {
    font-size: 12px;
    color: #64748b;
}
.summary-value {
    font-size: 18px;
    font-weight: 800;
    color: #0f172a;
    margin-top: 4px;
}
.combo-grid {
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    gap: 10px;
}
.combo-card {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 10px;
    padding: 10px 12px;
}
.combo-title {
    font-size: 13px;
    font-weight: 700;
    color: #1e293b;
    margin-bottom: 6px;
}
.combo-line {
    display: flex;
    justify-content: space-between;
    font-size: 12px;
    color: #475569;
    padding: 2px 0;
}
</style>
""", unsafe_allow_html=True)

if "show_floor_editor" not in st.session_state:
    st.session_state["show_floor_editor"] = False
if "custom_floor_heights" not in st.session_state:
    st.session_state["custom_floor_heights"] = []
if "spec_recommend_option" not in st.session_state:
    st.session_state["spec_recommend_option"] = "추천1"
if "grouped_signature" not in st.session_state:
    st.session_state["grouped_signature"] = None
if "glass_inputs" not in st.session_state:
    st.session_state["glass_inputs"] = [{"name": "유리1", "w": 2.0, "h": 2.0}]
if "rec1_price_inputs" not in st.session_state:
    st.session_state["rec1_price_inputs"] = _default_rec1_prices()
if "rec2_price_inputs" not in st.session_state:
    # 이전 버전의 combo_price_inputs가 있으면 추천2 초기값으로 이관
    legacy = st.session_state.get("combo_price_inputs", {})
    if isinstance(legacy, dict) and legacy:
        st.session_state["rec2_price_inputs"] = {
            row["spec"]: legacy.get(row["spec"], None)
            for row in RECOMMEND_PRICE_ROWS
        }
    else:
        st.session_state["rec2_price_inputs"] = _default_rec2_prices()

for i, g in enumerate(st.session_state.get("glass_inputs", []), start=1):
    if not isinstance(g, dict):
        st.session_state["glass_inputs"][i - 1] = {"name": f"유리{i}", "w": 2.0, "h": 2.0}
        continue
    g.setdefault("name", f"유리{i}")
    g.setdefault("w", 2.0)
    g.setdefault("h", 2.0)


def _build_grouped_floor_heights(num_floors, top_n, top_h, std_h, low_n, low_h):
    standard_count = int(num_floors) - int(top_n) - int(low_n)
    if standard_count < 0:
        return []
    return (
        [float(low_h)] * int(low_n)
        + [float(std_h)] * int(standard_count)
        + [float(top_h)] * int(top_n)
    )


manual_pdf_bytes, manual_pdf_name = _load_manual_pdf()

# ==========================================
# 1. 사이드바 (Input Data)
# ==========================================
with st.sidebar:
    st.markdown("<h2 style='font-size:1.2rem; font-weight:800; color:#0f172a;'>검토 조건 입력</h2>", unsafe_allow_html=True)
    if manual_pdf_bytes:
        st.download_button(
            label="사용 매뉴얼 다운로드",
            data=manual_pdf_bytes,
            file_name=manual_pdf_name,
            mime="application/pdf",
            key="download_manual_pdf",
        )

    st.markdown('<div class="sidebar-section-title">건물 정보</div>', unsafe_allow_html=True)

    num_floors = st.number_input("건물층수", value=25, min_value=3, max_value=100, step=1)

    c_top1, c_top2 = st.columns(2)
    with c_top1:
        top_n = st.number_input(
            "최고층 개수",
            value=1,
            min_value=0,
            max_value=200,
            step=1,
        )
    with c_top2:
        top_h = st.number_input(
            "최고층 층고 (m)",
            value=3.01,
            step=0.05,
            format="%.2f",
        )

    std_h = st.number_input(
        "기준층 층고 (m)",
        value=2.81,
        step=0.05,
        format="%.2f",
    )

    c_low1, c_low2 = st.columns(2)
    with c_low1:
        low_n = st.number_input(
            "최저층 개수",
            value=3,
            min_value=0,
            max_value=200,
            step=1,
        )
    with c_low2:
        low_h = st.number_input(
            "최저층 층고 (m)",
            value=3.01,
            step=0.05,
            format="%.2f",
        )

    grouped_sig = (
        int(num_floors),
        int(top_n),
        float(top_h),
        float(std_h),
        int(low_n),
        float(low_h),
    )
    if st.session_state.get("grouped_signature") != grouped_sig:
        st.session_state["grouped_signature"] = grouped_sig
        st.session_state["custom_floor_heights"] = []

    st.caption("생성된 층고를 확인하고 필요 시 직접 수정할 수 있습니다.")
    if st.button("층고 확인", use_container_width=True):
        st.session_state["show_floor_editor"] = True
    if st.session_state.get("custom_floor_heights"):
        st.caption(f"수정 적용됨: {len(st.session_state['custom_floor_heights'])}개 층")

    st.markdown("---")
    st.markdown('<div class="sidebar-group-title">풍하중 조건</div>', unsafe_allow_html=True)
    c_v, c_exp, c_pos = st.columns(3)
    with c_v:
        V0 = st.number_input("기본풍속 V₀ (m/s)", value=30.0, step=1.0, format="%.1f")
    with c_exp:
        exposure = st.selectbox("노풍도", ["A", "B", "C", "D"], index=2)
    with c_pos:
        position = st.selectbox("검토위치", ["중앙부", "모서리부"], index=0)

    st.markdown("---")
    st.markdown('<div class="sidebar-group-title">유리 크기</div>', unsafe_allow_html=True)
    b_add, b_del, _ = st.columns([1.2, 1.2, 3.6])
    with b_add:
        if st.button("추가", key="add_glass_size_top"):
            next_no = len(st.session_state["glass_inputs"]) + 1
            st.session_state["glass_inputs"].append({"name": f"유리{next_no}", "w": 2.0, "h": 2.0})
            st.rerun()
    with b_del:
        if st.button("삭제", key="delete_glass_size_top"):
            if len(st.session_state["glass_inputs"]) > 1:
                st.session_state["glass_inputs"].pop()
                st.rerun()
            else:
                st.warning("유리 크기는 최소 1개 이상 필요합니다.")

    glass_tabs = st.tabs([f"{idx + 1}" for idx in range(len(st.session_state["glass_inputs"]))])
    for idx, glass_tab in enumerate(glass_tabs):
        with glass_tab:
            _name = st.text_input(
                f"유리{idx + 1} 이름",
                value=str(st.session_state["glass_inputs"][idx].get("name", f"유리{idx + 1}")),
                key=f"glass_name_{idx}",
                placeholder="예: 1호 거실대창",
            )

            c_gw, c_gh = st.columns(2)
            with c_gw:
                _w = st.number_input(
                    "유리 폭 (m)",
                    value=float(st.session_state["glass_inputs"][idx].get("w", 2.0)),
                    step=0.1,
                    format="%.2f",
                    min_value=0.3,
                    key=f"glass_w_{idx}",
                )
            with c_gh:
                _h = st.number_input(
                    "유리 높이 (m)",
                    value=float(st.session_state["glass_inputs"][idx].get("h", 2.0)),
                    step=0.1,
                    format="%.2f",
                    min_value=0.3,
                    key=f"glass_h_{idx}",
                )
            st.session_state["glass_inputs"][idx] = {
                "name": _name.strip() if _name.strip() else f"유리{idx + 1}",
                "w": float(_w),
                "h": float(_h),
            }
    st.caption(f"총 {len(st.session_state['glass_inputs'])}개 유리 크기")

    st.markdown("---")
    st.markdown('<div class="sidebar-group-title">유리 Spec 추천 옵션</div>', unsafe_allow_html=True)
    st.caption("미입력 값은 후순위로 밀림 (단위: 원/m2)")

    rec_price_df = pd.DataFrame([
        {
            "구분": row["rank"],
            "재질 및 두께": row["spec"],
            "추천1(원)": st.session_state["rec1_price_inputs"].get(row["spec"], None),
            "추천2(원)": st.session_state["rec2_price_inputs"].get(row["spec"], None),
        }
        for row in RECOMMEND_PRICE_ROWS
    ])

    with st.form("rec_price_table_form", clear_on_submit=False):
        edited_price_df = st.data_editor(
            rec_price_df,
            use_container_width=False,
            hide_index=True,
            num_rows="fixed",
            key="rec_combo_price_editor",
            disabled=["구분", "재질 및 두께"],
            column_config={
                "구분": st.column_config.NumberColumn("구분", width="small"),
                "재질 및 두께": st.column_config.TextColumn("재질 및 두께", width="medium"),
                "추천1(원)": st.column_config.NumberColumn("추천1(원)", width="small", min_value=0, step=1000),
                "추천2(원)": st.column_config.NumberColumn("추천2(원)", width="small", min_value=0, step=1000),
            },
        )
        save_price_table = st.form_submit_button("단가표 저장", use_container_width=True)

    if save_price_table:
        for _, row in edited_price_df.iterrows():
            spec = str(row["재질 및 두께"])
            rec1_raw = row["추천1(원)"]
            rec2_raw = row["추천2(원)"]

            if pd.isna(rec1_raw):
                st.session_state["rec1_price_inputs"][spec] = None
            else:
                try:
                    st.session_state["rec1_price_inputs"][spec] = max(0.0, float(rec1_raw))
                except Exception:
                    st.session_state["rec1_price_inputs"][spec] = None

            if pd.isna(rec2_raw):
                st.session_state["rec2_price_inputs"][spec] = None
            else:
                try:
                    st.session_state["rec2_price_inputs"][spec] = max(0.0, float(rec2_raw))
                except Exception:
                    st.session_state["rec2_price_inputs"][spec] = None
        st.success("단가표 저장 완료")

    selected_option = st.radio(
        "추천 옵션 선택",
        ["추천1", "추천2", "추천3"],
        horizontal=True,
        key="recommend_option_picker",
        index=["추천1", "추천2", "추천3"].index(st.session_state.get("spec_recommend_option", "추천1")),
    )
    if selected_option == "추천1":
        st.caption("25년 평균 유리 단가")
    elif selected_option == "추천2":
        st.caption("유리 단가를 아는 경우, VE에 활용")
    if selected_option == "추천3":
        st.caption("구조 기준: 허용 풍압(LR)이 낮은 순으로 충족 조합을 추천")

    if st.button("선택 옵션 적용", use_container_width=True, key="apply_selected_option"):
        st.session_state["spec_recommend_option"] = selected_option
        st.success(f"{selected_option} 적용 완료")

    st.markdown("---")
    btn_calc = st.button("🔍 검토 실행", use_container_width=True, type="primary")

# ==========================================
# 2. 메인 영역 - 타이틀
# ==========================================
st.markdown("<h1 style='color:#1e293b; font-size:1.6rem; font-weight:800; margin-bottom:4px;'>단위세대 복층유리 재질/두께 검토</h1>", unsafe_allow_html=True)
st.markdown("""
<div style="font-size:13px; color:#64748b; margin-bottom:12px; line-height:1.6;">
    적용 기준: ASTM E1300-16 &nbsp;|&nbsp; 복층유리(IGU) &nbsp;|&nbsp; 5/6/8mm &nbsp;|&nbsp; 4변 지지 &nbsp;|&nbsp;
    AN: 일반유리, HS: 반강화유리, FT: 강화유리
</div>
""", unsafe_allow_html=True)

# ==========================================
# 2-1. 층고 직접입력 전용 화면
# ==========================================
if st.session_state.get("show_floor_editor", False):
    left_gap, popup_col, right_gap = st.columns([1.2, 1.6, 1.2], gap="large")
    with popup_col:
        st.markdown("""
        <div class="floor-popup-head">
            <div class="floor-popup-title">층고 확인</div>
            <div class="floor-popup-sub">층별 층고를 확인하고 직접 수정할 수 있습니다.</div>
        </div>
        """, unsafe_allow_html=True)

        floors = list(range(1, int(num_floors) + 1))
        grouped_heights = _build_grouped_floor_heights(num_floors, top_n, top_h, std_h, low_n, low_h)
        if st.session_state.get("custom_floor_heights") and len(st.session_state["custom_floor_heights"]) == int(num_floors):
            heights = st.session_state["custom_floor_heights"]
        else:
            heights = grouped_heights

        if len(heights) != int(num_floors):
            st.error("최고층 개수 + 최저층 개수가 건물층수를 초과했습니다. 입력값을 확인해 주세요.")
            if st.button("닫기", key="close_floor_editor_error"):
                st.session_state["show_floor_editor"] = False
                st.rerun()
            st.stop()

        edit_df = pd.DataFrame({"층": floors, "층고(m)": heights})
        edited_df = st.data_editor(
            edit_df,
            use_container_width=True,
            hide_index=True,
            disabled=["층"],
            num_rows="fixed",
            key="floor_height_editor",
            column_config={
                "층": st.column_config.NumberColumn("층", width="small"),
                "층고(m)": st.column_config.NumberColumn("층고(m)", width="medium", format="%.2f")
            }
        )
        b1, b2, _ = st.columns([1, 1, 3])
        with b1:
            if st.button("수정", type="primary"):
                _vals = edited_df["층고(m)"].tolist()
                if any(float(v) <= 0 for v in _vals):
                    st.error("층고는 0보다 커야 합니다.")
                else:
                    st.session_state["custom_floor_heights"] = [float(v) for v in _vals]
                    st.success("수정한 층고가 적용되었습니다.")
                    st.session_state["show_floor_editor"] = False
                    st.rerun()
        with b2:
            if st.button("닫기", key="close_floor_editor"):
                st.session_state["show_floor_editor"] = False
                st.rerun()

# ==========================================
# 3. 계산 실행
# ==========================================
if btn_calc:
    custom_heights = st.session_state.get("custom_floor_heights", [])
    use_custom_heights = isinstance(custom_heights, list) and len(custom_heights) == int(num_floors)
    mode = "individual" if use_custom_heights else "grouped"
    base_data = {
        "건물층수": str(num_floors),
        "최고층 층고 [m]": str(top_h),
        "기준층 층고 [m]": str(std_h),
        "최저층 층고 [m]": str(low_h),
        "기본풍속 (V0) [m/sec]": str(V0),
        "노풍도 (A, B, C, D)": exposure,
        "검토위치 (중앙부, 모서리부)": position,
        "입력방식": mode,
        "최고층 개수": str(top_n),
        "최저층 개수": str(low_n),
        "유리 Spec 추천 옵션": st.session_state.get("spec_recommend_option", "추천1"),
        "단가표": {
            "rec1_combo_prices": {
                spec: float(price)
                for spec, price in st.session_state.get("rec1_price_inputs", {}).items()
                if price is not None and str(price).strip() != ""
            },
            "rec2_combo_prices": {
                spec: float(price)
                for spec, price in st.session_state.get("rec2_price_inputs", {}).items()
                if price is not None and str(price).strip() != ""
            },
        },
    }
    if mode == "individual":
        base_data["층고목록"] = custom_heights

    calc_results = []
    for idx, glass in enumerate(st.session_state.get("glass_inputs", []), start=1):
        data = dict(base_data)
        data["단가표"] = dict(base_data.get("단가표", {}))
        data["유리 폭 (m)"] = str(glass.get("w", 2.0))
        data["유리 높이 (m)"] = str(glass.get("h", 2.0))
        outs = GlassCalculator.calculate_outputs(data, tab_type="brief")
        glass_name = str(glass.get("name", f"유리{idx}")).strip() or f"유리{idx}"
        calc_results.append({
            "index": idx,
            "label": glass_name,
            "glass_text": f"{glass.get('w', 2.0):.2f} × {glass.get('h', 2.0):.2f} m",
            "glass_size_short": f"{glass.get('w', 2.0):g}x{glass.get('h', 2.0):g}",
            "outs": outs,
            "data": data,
        })

    st.session_state["calc_results"] = calc_results

# ==========================================
# 4. 결과 표시 (대시보드: 좌=검토조건, 우=검토결과)
# ==========================================
if "calc_results" in st.session_state and st.session_state["calc_results"]:
    calc_results = st.session_state["calc_results"]
    first_item = calc_results[0]
    brief = first_item["outs"].get("brief", "")

    # 검토 조건 값 추출
    _data = first_item.get("data", {})
    _nf = _data.get("건물층수", "")
    _mode = _data.get("입력방식", "grouped")

    if _mode == "individual":
        _input_mode_label = "구간 입력(층고 수정 적용)"
        _th = "수정 반영 - '상세 계산 근거'에서 확인 가능"
        _sh = "수정 반영 - '상세 계산 근거'에서 확인 가능"
        _lh = "수정 반영 - '상세 계산 근거'에서 확인 가능"
    else:
        _input_mode_label = "구간 입력"
        _th = _data.get("최고층 층고 [m]", "") + " m" if _data.get("최고층 층고 [m]", "") else ""
        _sh = _data.get("기준층 층고 [m]", "") + " m" if _data.get("기준층 층고 [m]", "") else ""
        _lh = _data.get("최저층 층고 [m]", "") + " m" if _data.get("최저층 층고 [m]", "") else ""
    
    _v0 = _data.get("기본풍속 (V0) [m/sec]", "")
    _exp = _data.get("노풍도 (A, B, C, D)", "")
    _pos = _data.get("검토위치 (중앙부, 모서리부)", "")
    _glass_sizes = [item.get("glass_size_short", "") for item in calc_results if item.get("glass_size_short")]
    _glass_list = " / ".join(_glass_sizes) + " /" if _glass_sizes else "-"
    _rec_opt = _data.get("유리 Spec 추천 옵션", "추천1")

    # brief에서 건물높이 추출
    _H_line = [l for l in brief.splitlines() if "건물높이" in l]
    _H_val = _H_line[0].split(":")[1].strip() if _H_line else "-"

    st.markdown("---")
    left_col, right_col = st.columns([1, 1.4], gap="large")

    # ─── 왼쪽: 검토 조건 ───
    with left_col:
        cond_html = f'''
        <div class="panel-box">
            <div class="panel-title">검토 조건</div>
            <div class="cond-row"><span class="cond-label">입력방식</span><span class="cond-value">{_input_mode_label}</span></div>
            <div class="cond-row"><span class="cond-label">건물층수</span><span class="cond-value">{_nf}층</span></div>
            <div class="cond-row"><span class="cond-label">최고층 층고</span><span class="cond-value">{_th}</span></div>
            <div class="cond-row"><span class="cond-label">기준층 층고</span><span class="cond-value">{_sh}</span></div>
            <div class="cond-row"><span class="cond-label">최저층 층고</span><span class="cond-value">{_lh}</span></div>
            <div class="cond-row"><span class="cond-label">건물높이 (H)</span><span class="cond-value">{_H_val}</span></div>
            <div class="cond-row"><span class="cond-label">기본풍속 (V₀)</span><span class="cond-value">{_v0} m/s</span></div>
            <div class="cond-row"><span class="cond-label">노풍도</span><span class="cond-value">{_exp}</span></div>
            <div class="cond-row"><span class="cond-label">검토위치</span><span class="cond-value">{_pos}</span></div>
            <div class="cond-row"><span class="cond-label">유리크기</span><span class="cond-value">{_glass_list}</span></div>
            <div class="cond-row"><span class="cond-label">유리 Spec 추천 옵션</span><span class="cond-value">{_rec_opt}</span></div>
            <div class="cond-row"><span class="cond-label">처짐기준</span><span class="cond-value">L/60</span></div>
        </div>
        '''
        st.markdown(cond_html, unsafe_allow_html=True)

    # ─── 오른쪽: 검토 결과 ───
    with right_col:
        result_tabs = st.tabs([f"{item['label']}" for item in calc_results])
        group_re = re.compile(r"^\d+\s*층\s*~\s*\d+\s*층\s*:")
        for tab, item in zip(result_tabs, calc_results):
            with tab:
                item_brief = item.get("outs", {}).get("brief", "")
                result_html_lines = []
                in_result = False
                for raw in item_brief.splitlines():
                    line = raw.strip()
                    if not line:
                        continue
                    if line.startswith("▶ 프로그램 개요") or line.startswith("▶ 검토 조건"):
                        in_result = False
                        continue
                    if line.startswith("- ") and not in_result:
                        continue
                    if line.startswith("※") and not in_result:
                        continue
                    if "▶ 검토 결과" in line:
                        in_result = True
                        result_html_lines.append(f'<div class="result-header">{line}</div>')
                        continue
                    if in_result:
                        if group_re.match(line):
                            result_html_lines.append(f'<div class="result-group">{line}</div>')
                        elif line.startswith("▶ "):
                            in_result = False
                        else:
                            result_html_lines.append(f'<div class="result-normal">{line}</div>')

                # 각 탭별로 해당 유리의 summary 생성
                item_data = item.get("data", {})
                rec_option = item_data.get("유리 Spec 추천 옵션", "추천1")
                
                item_test_lines = []
                # 추천3으로 검토시 유리면적 및 예상금액 표시 안함 (단가 입력 불가)
                if rec_option != "추천3":
                    item_summary = _build_area_cost_summary_for_item(item, calc_results)
                    if item_summary.get("rows"):
                        item_test_lines.append('<div class="result-header" style="margin-top:10px;">▶ 유리 면적 및 예상금액</div>')
                        for i, r in enumerate(item_summary.get("rows", []), start=1):
                            spec = r.get("spec", "-")
                            area = float(r.get("area", 0.0))
                            price = r.get("price", None)
                            amount = r.get("amount", None)
                            price_txt = f"{int(price):,}원/m2" if price is not None else "미입력"
                            amount_txt = f"{int(amount):,}원" if amount is not None else "확인 불가"
                            item_test_lines.append(f'<div class="result-normal">{i}. {spec} | 면적 {area:.2f}m2 | 단가 {price_txt} | 합계 {amount_txt}</div>')
                        item_test_lines.append(f'<div class="result-normal">총합계 면적: {float(item_summary.get("total_area", 0.0)):.2f}m2</div>')
                        if item_summary.get("missing_specs"):
                            item_test_lines.append('<div class="result-normal">해당 조합의 단가 입력 후 확인 가능</div>')
                        else:
                            item_test_lines.append(f'<div class="result-normal">예상금액: {int(float(item_summary.get("total_amount", 0.0))):,}원</div>')

                if item_test_lines:
                    result_html_lines.extend(item_test_lines)

                result_body = "\n".join(result_html_lines) if result_html_lines else '<div class="result-normal" style="color:#94a3b8;">검토 결과가 없습니다.</div>'
                st.markdown(f'''
                <div class="panel-box">
                    <div class="panel-title">검토 결과</div>
                    {result_body}
                </div>
                ''', unsafe_allow_html=True)

    # ─── 다운로드 버튼 ───
    col1, col2, col_gap = st.columns([1.5, 1.5, 3])
    with col1:
        excel_bytes = _generate_excel(calc_results)
        if excel_bytes:
            st.download_button(
                label="검토결과 내보내기(엑셀)",
                data=excel_bytes,
                file_name=f"유리검토_결과정리_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    
    with col2:
        merged_detail_parts = []
        for item in calc_results:
            merged_detail_parts.append(f"# {item['index']} | {item['label']} ({item['glass_text']})")
            merged_detail_parts.append(item.get("outs", {}).get("detail", ""))
        pdf_bytes = _generate_pdf("\n\n".join(merged_detail_parts))
        if pdf_bytes:
            st.download_button(
                label="계산근거 내보내기(PDF)",
                data=pdf_bytes,
                file_name=f"유리검토_다중계산근거_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf",
            )
        else:
            st.warning("PDF 폰트를 찾을 수 없어 다운로드를 사용할 수 없습니다.")

    # ─── 상세 계산 근거 (토글) ───
    with st.expander("📋 상세 계산 근거 보기", expanded=False):
        detail_tabs = st.tabs([f"{item['label']}" for item in calc_results])
        for tab, item in zip(detail_tabs, calc_results):
            with tab:
                detail = item.get("outs", {}).get("detail", "")
                sections = detail.split("# 상세(계산 근거)")
                if len(sections) <= 1:
                    st.caption("상세 계산 근거가 없습니다.")
                    continue

                detail_body = sections[1]
                parts = detail_body.split("# 조합 근거(층별)")
                table_part = parts[0].strip() if len(parts) > 0 else ""
                combo_part = parts[1].strip() if len(parts) > 1 else ""

                st.markdown("#### 상세(계산 근거)")
                table_lines = [l for l in table_part.splitlines() if l.strip()]
                if table_lines:
                    headers = [h.strip() for h in table_lines[0].split("|")]
                    html_table = '<table class="detail-table"><thead><tr>'
                    for h in headers:
                        html_table += f'<th>{h}</th>'
                    html_table += '</tr></thead><tbody>'
                    for row_line in table_lines[1:]:
                        cols = [c.strip() for c in row_line.split("|")]
                        html_table += '<tr>'
                        for col_val in cols:
                            html_table += f'<td>{col_val}</td>'
                        html_table += '</tr>'
                    html_table += '</tbody></table>'
                    st.markdown(html_table, unsafe_allow_html=True)

                if combo_part:
                    st.markdown("#### 조합 근거(층별)")
                    floor_re = re.compile(r"^(\d+)층\s*\(설계풍압\s*([\d.]+)\s*kN")
                    current_floor_lines = []
                    current_title = ""

                    def render_combo_block(title, lines):
                        if not title:
                            return
                        has_price = any(len(l.strip("*").split("|")) >= 3 for l in lines if l.strip())

                        html = f'<details><summary style="font-weight:700; cursor:pointer; margin:4px 0;">{title}</summary>'
                        if has_price:
                            html += '<table class="detail-table"><thead><tr><th>조합</th><th>판정</th><th>단가 계산</th></tr></thead><tbody>'
                        else:
                            html += '<table class="detail-table"><thead><tr><th>조합</th><th>판정</th></tr></thead><tbody>'

                        for cl in lines:
                            is_picked = cl.startswith("**") and cl.endswith("**")
                            cl_clean = cl.strip("*")
                            parts_c = cl_clean.split("|")
                            spec = parts_c[0].strip() if len(parts_c) > 0 else cl_clean
                            ok = parts_c[1].strip() if len(parts_c) > 1 else ""
                            price_calc = parts_c[2].strip() if len(parts_c) > 2 else ""
                            css_class = "combo-picked" if is_picked else ""
                            ok_class = "combo-ok" if ok == "OK" else "combo-ng"

                            if has_price:
                                html += f'<tr class="{css_class}"><td>{spec}</td><td class="{ok_class}">{ok}</td><td>{price_calc}</td></tr>'
                            else:
                                html += f'<tr class="{css_class}"><td>{spec}</td><td class="{ok_class}">{ok}</td></tr>'
                        html += '</tbody></table></details>'
                        st.markdown(html, unsafe_allow_html=True)

                    for cl in combo_part.splitlines():
                        cl = cl.strip()
                        if not cl:
                            continue
                        m = floor_re.match(cl)
                        if m:
                            render_combo_block(current_title, current_floor_lines)
                            current_title = cl
                            current_floor_lines = []
                        else:
                            current_floor_lines.append(cl)
                    render_combo_block(current_title, current_floor_lines)
