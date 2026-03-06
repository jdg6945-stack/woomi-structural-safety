import os
import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import platform
import io
import json
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================
# 1. 환경 설정 및 스타일링
# ==========================================
def set_env():
    if platform.system() == 'Windows': plt.rc('font', family='Malgun Gothic')
    elif platform.system() == 'Darwin': plt.rc('font', family='AppleGothic')
    else: plt.rc('font', family='NanumGothic')
    plt.rcParams['axes.unicode_minus'] = False
    try:
        st.set_page_config(layout="wide", page_title="PRD 축력 검토")
    except Exception:
        pass
    
    st.markdown("""
        <style>
        /* Modern Premium Grey Style */
        .stApp { background-color: #ffffff; }
        
        .report-header { 
            background: linear-gradient(135deg, #1e293b 0%, #475569 100%); 
            color: white; 
            padding: 25px 30px; 
            border-radius: 12px; 
            margin-bottom: 25px; 
            box-shadow: 0 10px 15px -3px rgba(0, 0, 0, 0.1); 
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .report-section { background: white; padding: 25px; border-radius: 12px; border: 1px solid #e2e8f0; margin-bottom: 20px; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05); }
        .section-title { color: #1e293b; border-left: 6px solid #475569; padding-left: 15px; margin-bottom: 20px; font-size: 1.3rem; font-weight: 800; }
        
        .metric-card { text-align: center; padding: 20px; background: #ffffff; border-radius: 12px; border: 1px solid #e2e8f0; box-shadow: 0 2px 4px rgba(0,0,0,0.05); }
        .metric-label { font-size: 0.9rem; color: #64748b; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em; }
        .metric-value { font-size: 1.8rem; color: #0f172a; font-weight: 900; margin-top: 5px; }
        
        .ok-banner { background-color: #f0fdf4; color: #166534; padding: 20px; border-radius: 10px; border: 1px solid #bbf7d0; font-weight: 700; font-size: 1.2rem; text-align: center; }
        .ng-banner { background-color: #fef2f2; color: #991b1b; padding: 20px; border-radius: 10px; border: 1px solid #fecaca; font-weight: 700; font-size: 1.2rem; text-align: center; }
        
        [data-testid="stSidebar"] { background-color: #f8fafc; border-right: 1px solid #e2e8f0; }
        .sidebar-title { color: #334155; font-size: 1.2rem; font-weight: 800; margin-bottom: 10px; padding: 10px 0; border-bottom: 2px solid #e2e8f0; }
        
        div.stButton > button {
            width: 100%;
            height: 3rem;
            background-color: #f1f5f9 !important;
            color: #334155 !important;
            border-radius: 8px !important;
            font-weight: 700 !important;
            transition: all 0.2s ease;
            border: 1px solid #e2e8f0 !important;
        }
        div.stButton > button:hover {
            transform: translateY(-2px);
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.05);
            background-color: #e2e8f0 !important;
            border-color: #cbd5e1 !important;
        }
        
        .stTabs [data-baseweb="tab-list"] { gap: 10px; }
        .stTabs [data-baseweb="tab"] { 
            background-color: white; 
            border-radius: 8px 8px 0 0; 
            padding: 10px 20px; 
            border: 1px solid #e2e8f0; 
            font-weight: 600;
            color: #64748b;
        }
        .stTabs [aria-selected="true"] { background-color: #f1f5f9 !important; color: #1e293b !important; border-bottom: 2px solid #475569 !important; }
        
        /* Input Field Border Styling */
        div[data-testid="stNumberInput"] [data-baseweb="input"],
        div[data-testid="stSelectbox"] [data-baseweb="select"],
        div[data-testid="stTextInput"] [data-baseweb="input"] {
            border: 1px solid #cbd5e1 !important;
            border-radius: 4px !important;
            background-color: #ffffff !important;
            min-height: 40px !important;
        }
        
        /* Default: Hide +/- buttons */
        div.hide-btns [data-testid="stNumberInput"] button {
            display: none !important;
        }
        
        /* Specific targeting for better control if needed */
        div[data-testid="stNumberInput"] button {
            display: inline-flex; /* Standard buttons restored */
        }
        
        /* Remove internal borders that cause partial visibility */
        div[data-testid="stNumberInput"] > div {
            border: none !important;
        }
        
        /* Ensure the input text is centered and has padding */
        div[data-testid="stNumberInput"] input {
            padding: 8px 12px !important;
        }
        
        /* Focused status */
        div[data-testid="stNumberInput"] [data-baseweb="input"]:focus-within, 
        div[data-testid="stSelectbox"] [data-baseweb="select"]:focus-within {
            border-color: #475569 !important;
            box-shadow: 0 0 0 1px rgba(71, 85, 105, 0.2) !important;
        }
        /* Ultra-compact widget spacing */
        div[data-testid="stSidebar"] [data-testid="stVerticalBlock"] > div {
            margin-bottom: -24px !important;
        }
        div[data-testid="stSidebar"] label {
            font-size: 0.72rem !important;
            font-weight: 700 !important;
            margin-bottom: -18px !important;
            color: #475569 !important;
        }
        /* Minimal height for inputs and components */
        div[data-testid="stSidebar"] div[data-baseweb="input"] > div {
            min-height: 24px !important;
            height: 24px !important;
            padding: 0px 6px !important;
        }
        div[data-testid="stSidebar"] input {
            font-size: 0.8rem !important;
            padding: 0px !important;
        }
        /* Tight selectboxes */
        div[data-testid="stSidebar"] div[data-baseweb="select"] > div {
            min-height: 24px !important;
            height: 24px !important;
            font-size: 0.75rem !important;
            padding: 0px 4px !important;
        }
        /* Section titles - Larger and Bold */
        .sidebar-title {
            font-size: 1.15rem !important;
            margin-bottom: 0px !important;
            margin-top: 10px !important;
            font-weight: 800 !important;
            color: #1e293b !important;
        }
        /* Remove borders from containers and tabs in sidebar */
        div[data-testid="stSidebar"] [data-testid="stVerticalBlock"] div[data-testid="stExpander"] {
            border: none !important;
        }
        div[data-testid="stSidebar"] [data-baseweb="tab-list"] {
            border-bottom: none !important;
        }
        div[data-testid="stSidebar"] [data-baseweb="tab-list"] button {
            border-bottom: none !important;
        }
        div[data-testid="stSidebar"] div[data-testid="stVerticalBlock"] > div[style*="border"] {
            border: none !important;
        }
        /* Extra tightness for ultra-compact after removing borders */
        div[data-testid="stSidebar"] div.stElementContainer {
            margin-bottom: -22px !important;
        }
        div[data-testid="stSidebar"] [data-testid="stForm"] {
            padding: 5px !important;
        }
        /* Reduce gap between columns in sidebar */
        div[data-testid="stSidebar"] [data-testid="column"] {
            gap: 0.5rem !important;
        }
        /* Data Editor Header Wrap Fix */
    [data-testid="stDataEditor"] [data-testid="stHeader"] {
        white-space: pre-line !important;
        height: auto !important;
        min-height: 40px !important;
        display: flex !important;
        align-items: center !important;
    }
    [data-testid="stDataEditor"] div[role="columnheader"] {
        padding-top: 5px !important;
        padding-bottom: 5px !important;
    }
    </style>
        """, unsafe_allow_html=True)

UNIT_C, UNIT_S = 24.0, 78.5

# ==========================================
# 2. 계산 엔진
# ==========================================
class PRDEngine:
    @staticmethod
    def calculate_floor(row, is_top=False, is_bottom=False, override_ll=None):
        area = row["X분담(m)"] * row["Y분담(m)"]
        d_logs = []
        
        t_m = row["Slab(m)"]
        w_slab = t_m * UNIT_C * area
        d_logs.append(f"슬래브: {area:.2f}m2 x {t_m:.3f}m x {UNIT_C} = {w_slab:.2f} kN")
        
        w_beam = 0
        if not is_bottom:
            b_list = row.get('b_list', [])
            for i, b in enumerate(b_list):
                b_type = b.get('type', 'RC')
                if b_type == "RC":
                    eff_bh = (b['bh']/1000 - row["Slab(m)"])
                    v_b = (b['bw']/1000) * eff_bh * b['bl'] * UNIT_C
                    d_logs.append(f"보#{i+1}(RC): {b['bw']/1000:.3f}m(B) x {eff_bh:.3f}m(H) x {b['bl']:.2f}m(L) x {UNIT_C} = {v_b:.2f} kN")
                else:
                    s_area = (b['h']*b['tw'] + 2*b['b']*b['tf'] - 2*b['tw']*b['tf']) / 10**6
                    v_b = s_area * b['bl'] * UNIT_S
                    d_logs.append(f"보#{i+1}(S): {s_area:.6f}m2 x {b['bl']}m x {UNIT_S} = {v_b:.2f} kN")
                w_beam += v_b
        
        w_col = 0
        is_roof = (row.get("층이름") == "Roof")
        if not is_bottom and not is_roof:
            cv = row['c_vals']
            c_unit = UNIT_C if row['c_type'] == "RC" else UNIT_S
            if row['c_type'] == "RC":
                ca = (cv['cx']/1000 * cv['cy']/1000)
            elif row['c_type'] == "철골(ㅁ)":
                ca = (cv['h']*cv['b'] - (cv['h']-2*cv['t'])*(cv['b']-2*cv['t']))/10**6
            else:
                ca = (cv['h']*cv['tw']+2*cv['b']*cv['tf']-2*cv['tw']*cv['tf'])/10**6
            w_col = ca * row["층고(m)"] * c_unit
            d_logs.append(f"기둥({row['c_type']}): {ca:.6f}m2 x {row['층고(m)']:.2f}m(H) x {c_unit} = {w_col:.2f} kN")
        elif is_roof:
            d_logs.append("기둥: 지붕층 상부 수직부재 없음 (0.00 kN)")
        
        # Determine which LL to use
        active_ll = float(override_ll) if override_ll is not None else float(row["LL(kN/m2)"])
        
        add_dl = row.get("추가하중(kN/m2)", 0.0) * area
        d_logs.append(f"기타하중(DL): {row.get('추가하중(kN/m2)', 0.0):.2f}kN/m2 x {area:.2f}m2 = {add_dl:.2f} kN")
            
        dl_sum = w_slab + w_beam + w_col + add_dl
        ll_total = active_ll * area
        d_logs.append(f"활하중(LL): {active_ll:.2f}kN/m2 x {area:.2f}m2 = {ll_total:.2f} kN")
        
        f_total = dl_sum + ll_total 
        
        return f_total, {"DL_Comp": w_slab + w_beam + w_col, "DL_Add": add_dl, "DL_Sum": dl_sum, "LL_Sum": ll_total, "Slab_Weight": w_slab, "Beam_Weight": w_beam, "Column_Weight": w_col, "Logs": d_logs}

    @classmethod
    def run_analysis(cls, df_input, prd_limit):
        res = pd.DataFrame(df_input)
        all_logs = {}
        
        # 1. 먼저 모든 층의 고유 하중(Basic Usage)을 계산하여 저장
        for i in range(len(res)):
            # 하중 계산 시 기본적으로 입력된 LL 사용
            usage, logs = cls.calculate_floor(res.iloc[i], is_top=(i==0), is_bottom=(i==len(res)-1))
            res.loc[i, "Floor_Usage_Basic"] = usage
            res.loc[i, "DL(kN)"] = logs["DL_Sum"]
            res.loc[i, "추가하중(kN)"] = logs["DL_Add"]
            res.loc[i, "LL(kN)"] = logs["LL_Sum"]
            
            # 부가 정보 (면적, 보 무게 등)
            try:
                area = float(res.loc[i, "X분담(m)"]) * float(res.loc[i, "Y분담(m)"])
            except:
                area = 0.0
                
            res.loc[i, "면적"] = round(area, 2)
            # Individual Floor Weights (Transparency)
            res.loc[i, "슬래브 무게(kN)"] = round(float(logs.get("Slab_Weight", 0)), 1)
            res.loc[i, "보 무게(kN)"] = round(float(logs.get("Beam_Weight", 0)), 1)
            res.loc[i, "기둥 무게(kN)"] = round(float(logs.get("Column_Weight", 0)), 1)
            
            # Individual Floor Resultant Loads
            dl_floor = float(logs.get("DL_Sum", 0))
            ll_floor = float(logs.get("LL_Sum", 0))
            res.loc[i, "단일 DL(kN)"] = round(dl_floor, 1)
            res.loc[i, "단일 LL(kN)"] = round(ll_floor, 1)
            res.loc[i, "단일 DL+LL(kN)"] = round(dl_floor + ll_floor, 1)
            res.loc[i, "단일 1.2D+1.6L(kN)"] = round(1.2*dl_floor + 1.6*ll_floor, 1)
            res.loc[i, "단일 1.2D+1.28L(kN)"] = round(1.2*dl_floor + 1.28*ll_floor, 1)
            
            all_logs[res.loc[i, "층이름"]] = logs

        # 2. 시공 가능 층수 판단 (Bottom-up Iteration)
        # 최하층(Foundation)인 마지막 인덱스는 무게 계산에서 제외됨
        res["판정"] = "NG"
        res["DL+LL"] = 0.0
        res["1.2D + 1.6L"] = 0.0
        res["1.2D + 1.28L"] = 0.0
        res["필요_PRD_축력"] = 0.0
        last_ok_idx = -1
        
        # 하부층부터 하나씩 위로 시나리오별 총하중 시뮬레이션
        for target_idx in range(len(res)-2, -1, -1):
            cur_usage = res.iloc[target_idx]["Floor_Usage_Basic"]
            
            sub_load = 0
            if target_idx < len(res)-2:
                sub_load = res.iloc[target_idx+1 : len(res)-1]["Floor_Usage_Basic"].sum()
                
            total_at_this_stage = cur_usage + sub_load
            res.loc[target_idx, "필요_PRD_축력"] = total_at_this_stage
            
            if total_at_this_stage <= prd_limit:
                # 이 층까지는 시공 가능함! 더 위로 검색 계속
                last_ok_idx = target_idx
        
        # 3. 판정 업데이트 및 결과값 세팅
        if last_ok_idx != -1:
            res.loc[last_ok_idx:, "판정"] = "OK"
            
            # Re-calculate loads for the final report based on the input LL
            res["DL+LL"] = 0.0
            res["1.2D + 1.6L"] = 0.0
            res["1.2D + 1.28L"] = 0.0
            
            working_range = res.iloc[last_ok_idx : len(res)-1]
            
            if not working_range.empty:
                temp_dls = []
                temp_lls = []
                
                for idx in working_range.index:
                    actual_ll = float(res.loc[idx, "LL(kN/m2)"])
                    
                    # Get fresh DL/LL for this floor with the correct active LL
                    _, f_logs = cls.calculate_floor(res.loc[idx], is_top=(idx==0), is_bottom=(idx==len(res)-1), override_ll=actual_ll)
                    
                    # Update snapshot results
                    res.loc[idx, "LL(kN)"] = f_logs["LL_Sum"]
                    temp_dls.append(f_logs["DL_Sum"])
                    temp_lls.append(f_logs["LL_Sum"])
                
                # Cumulative Sum (Top downwards)
                for i_pos, idx in enumerate(working_range.index):
                    acc_dl = sum(temp_dls[:i_pos+1])
                    acc_ll = sum(temp_lls[:i_pos+1])
                    # Final Results (Cumulative)
                    res.loc[idx, "DL+LL"] = round(acc_dl + acc_ll, 1)
                    res.loc[idx, "1.2D + 1.6L"] = round(1.2*acc_dl + 1.6*acc_ll, 1)
                    res.loc[idx, "1.2D + 1.28L"] = round(1.2*acc_dl + 1.28*acc_ll, 1)

        return res, (res.loc[last_ok_idx, "층이름"] if last_ok_idx != -1 else "시공 불가"), all_logs

# ==========================================
# 3. UI 헬퍼 함수
# ==========================================
def member_input_form_compact(key_p, defaults=None):
    with st.container():
        st.markdown('<div class="hide-btns">', unsafe_allow_html=True)
        # 기본 제원 (층고, 슬래브, 시공하중 등) - 3열 배치
        c1, c2, c3 = st.columns([1,1,1])
        h = c1.number_input("층고(m)", value=defaults['h'] if defaults else 3.5, step=0.1, key=f"{key_p}_h")
        s = c2.number_input("슬래브(m)", value=defaults['s'] if defaults else 0.2, step=0.01, key=f"{key_p}_s")
        ll = c3.number_input("시공하중(kN/m²)", value=defaults['ll'] if defaults else 1.0, step=0.1, key=f"{key_p}_ll")
        
        c1, c2, c3 = st.columns([1,1,1])
        x = c1.number_input("X분담(m)", value=defaults['x'] if defaults else 8.2, step=0.1, key=f"{key_p}_x")
        y = c2.number_input("Y분담(m)", value=defaults['y'] if defaults else 8.2, step=0.1, key=f"{key_p}_y")
        add_ll = c3.number_input("기타 DL(kN/m²)", value=defaults['add_ll'] if defaults else (0.0 if key_p == "F" else 0.0), step=0.1, key=f"{key_p}_add")
        st.markdown('</div>', unsafe_allow_html=True)

        # 기둥 정보 (테두리 적용)
        with st.container(border=True):
            st.markdown("<p style='font-size:0.85rem; font-weight:bold; margin-bottom:5px;'>기둥 정보</p>", unsafe_allow_html=True)
            st.markdown('<div class="hide-btns">', unsafe_allow_html=True)
            ct = st.selectbox("기둥 유형", ["RC", "철골(ㅁ)", "철골(H)"], 
                              index=0 if not defaults else ["RC", "철골(ㅁ)", "철골(H)"].index(defaults['ct']), 
                              key=f"{key_p}_ct")
            
            cv = defaults['cv'].copy() if defaults else {"cx":800, "cy":800, "h":500, "b":500, "t":25, "tw":16, "tf":25}
            
            # 4열 레이아웃 적용
            cols = st.columns(4)
            if ct == "RC":
                cv['cx'] = cols[0].number_input("폭(mm)", value=int(cv['cx']), step=10, key=f"{key_p}_cx")
                cv['cy'] = cols[1].number_input("깊이(mm)", value=int(cv['cy']), step=10, key=f"{key_p}_cy")
            elif ct == "철골(ㅁ)":
                cv['h'] = cols[0].number_input("H(mm)", value=int(cv['h']), step=10, key=f"{key_p}_ch")
                cv['b'] = cols[1].number_input("B(mm)", value=int(cv['b']), step=10, key=f"{key_p}_cb")
                cv['t'] = cols[2].number_input("t(mm)", value=int(cv['t']), step=1, key=f"{key_p}_cthick")
            else: # 철골(H)
                cv['h'] = cols[0].number_input("H(mm)", value=int(cv['h']), step=10, key=f"{key_p}_chh")
                cv['b'] = cols[1].number_input("B(mm)", value=int(cv['b']), step=10, key=f"{key_p}_cbb")
                cv['tw'] = cols[2].number_input("tw(mm)", value=int(cv['tw']), step=1, key=f"{key_p}_ctw")
                cv['tf'] = cols[3].number_input("tf(mm)", value=int(cv['tf']), step=1, key=f"{key_p}_ctf")
            st.markdown('</div>', unsafe_allow_html=True)

        # 보 정보 (테두리 적용)
        with st.container(border=True):
            st.markdown("<p style='font-size:0.85rem; font-weight:bold; margin-bottom:5px;'>보 정보</p>", unsafe_allow_html=True)
            d_bl = defaults['b_list'] if defaults else [{"type": "RC", "bl": 8.2, "bw": 500, "bh": 700}]
            # Beam count RESTORES +/- buttons (no hide-btns wrapper)
            b_cnt = st.number_input("보 수량", min_value=0, max_value=8, value=len(d_bl), key=f"{key_p}_bcnt")
            
            b_list = []
            if b_cnt > 0:
                st.markdown('<div class="hide-btns">', unsafe_allow_html=True)
                b_tabs = st.tabs([f"#{i+1}" for i in range(int(b_cnt))])
                # Fallback beam info if d_bl is empty
                fallback_b = {"type": "RC", "bl": 8.2, "bw": 500, "bh": 700}
                for i, current_tab in enumerate(b_tabs):
                    prev_b = d_bl[i] if i < len(d_bl) else (d_bl[0] if len(d_bl) > 0 else fallback_b)
                    with current_tab:
                        bcols = st.columns(4)
                        bt = bcols[0].selectbox(f"종류", ["RC", "철골(H)"], index=0 if prev_b.get('type')=="RC" else 1, key=f"{key_p}_bt_{i}")
                        b_len = bcols[1].number_input(f"L(m)", value=float(prev_b['bl']), step=0.1, key=f"{key_p}_blen_{i}")
                        
                        if bt == "RC":
                            bw = bcols[2].number_input(f"W(mm)", value=int(prev_b.get('bw', 500)), step=10, key=f"{key_p}_bw_{i}")
                            bh = bcols[3].number_input(f"H(mm)", value=int(prev_b.get('bh', 700)), step=10, key=f"{key_p}_bh_{i}")
                            b_list.append({"type": "RC", "bl": b_len, "bw": bw, "bh": bh})
                        else: # 철골(H) 보
                            bh_s = bcols[2].number_input(f"H(mm)", value=int(prev_b.get('h', 500)), step=10, key=f"{key_p}_bhs_{i}")
                            bb_s = bcols[3].number_input(f"B(mm)", value=int(prev_b.get('b', 300)), step=10, key=f"{key_p}_bbs_{i}")
                            
                            # 철골 보의 두께 정보는 추가 행(4열)으로 배치
                            bcols2 = st.columns(4)
                            btw_s = bcols2[0].number_input(f"tw(mm)", value=int(prev_b.get('tw', 10)), step=1, key=f"{key_p}_btws_{i}")
                            btf_s = bcols2[1].number_input(f"tf(mm)", value=int(prev_b.get('tf', 16)), step=1, key=f"{key_p}_btfs_{i}")
                            b_list.append({"type": "Steel", "bl": b_len, "h": bh_s, "b": bb_s, "tw": btw_s, "tf": btf_s})
                st.markdown('</div>', unsafe_allow_html=True)
                    
    return {"h":h, "x":x, "y":y, "s":s, "ll":ll, "add_ll":add_ll, "ct":ct, "cv":cv, "b_list":b_list}

def draw_professional_diagram(res, limit):
    fig, ax = plt.subplots(figsize=(7, 9))
    
    # 배경 (Grey Theme)
    ax.set_facecolor('#f8fafc')
    
    y_coords = range(len(res))
    labels = res['층이름'].tolist()
    
    for i, row in res.iterrows():
        y = len(res) - 1 - i
        is_ok = row["판정"] == "OK"
        
        # 층 바닥 라인 - Grey Tone
        color = "#1e293b" if is_ok else "#e2e8f0"
        alpha = 1.0 if is_ok else 0.5
        ax.hlines(y, 0, 100, colors=color, linewidth=4, alpha=alpha, zorder=3)
        
        # 라벨링
        ax.text(-5, y, labels[i], ha='right', va='center', fontsize=11, fontweight='bold', color='#1e293b')
        
        # 하중 수치 표시
        if is_ok:
            load_text = f"{row['DL+LL']:,.0f} kN"
            ax.text(105, y, load_text, ha='left', va='center', fontsize=10, fontweight='bold', color='#475569')

    # 기초 표시
    ax.fill_between([-15, 120], [-1.5, -1.5], [0, 0], color='#cbd5e1', alpha=0.2, hatch='///', zorder=1)
    ax.text(50, -0.8, "기초", ha='center', va='center', fontsize=10, fontweight='bold', color='#64748b')

    ax.set_xlim(-20, 130)
    ax.set_ylim(-1.5, len(res))
    
    # Hide all axis elements
    for spine in ax.spines.values(): spine.set_visible(False)
    ax.set_yticks([])
    ax.set_xticks([])
    
    plt.tight_layout()
    return fig

# ==========================================
# 4. Main App
# ==========================================
def main():
    set_env()
    
    if "prd_df" not in st.session_state: st.session_state.prd_df = None
    if "calc_results" not in st.session_state: st.session_state.calc_results = None
    
    # Early unpack results if available for sidebar Export
    res, last_ok, logs = None, None, None
    if st.session_state.calc_results:
        res, last_ok, logs = st.session_state.calc_results

    # --- SIDEBAR INPUTS ---
    with st.sidebar:
        st.markdown("<div class='sidebar-title'>1 단계: 프로젝트 설정</div>", unsafe_allow_html=True)
        with st.container():
            sc1, sc2 = st.columns(2)
            f_cnt = sc1.number_input("지상(F)", min_value=1, value=10)
            b_cnt = sc2.number_input("지하(B)", min_value=1, value=7)
            
        with st.container():
            sc3, sc4 = st.columns(2)
            limit = sc3.number_input("PRD축력(kN)", value=8000, step=100)
            target_f = sc4.number_input("목표층(지상)", min_value=1, value=int(f_cnt))
        
        if st.session_state.prd_df:
            st.markdown("---")
            if st.button("모든 데이터 초기화"):
                st.session_state.prd_df = None
                st.session_state.calc_results = None
                if 'run_analysis' in st.session_state:
                    st.session_state.run_analysis = False
                st.rerun()

        # Step 2: Base Config - Hide if data is already loaded to avoid confusion
        if st.session_state.prd_df is None:
            st.markdown("<div class='sidebar-title'>2 단계: 기본 제원 설정</div>", unsafe_allow_html=True)
            tab_f, tab_b = st.tabs(["지상층 (F)", "지하층 (B)"])
            with tab_f: f_set = member_input_form_compact("F")
            with tab_b: b_set = member_input_form_compact("B")
            
            st.write("")
            if st.button("기본 데이터 적용", type="primary"):
                st.session_state['run_analysis'] = True
        else:
            # Provide dummy sets to keep script running if needed elsewhere (though new_data logic is skipped)
            f_set = b_set = None


    # --- MAIN DASHBOARD ---
    if st.session_state.get('calc_results') is not None:
        if "last_limit" not in st.session_state:
            st.session_state.last_limit = limit
            
        if st.session_state.last_limit != limit:
            res, last_ok, logs = PRDEngine.run_analysis(st.session_state.prd_df, limit)
            st.session_state.calc_results = (res, last_ok, logs)
            st.session_state.last_limit = limit
            st.rerun()
            
        # Re-unpack in case it was updated
        res, last_ok, logs = st.session_state.calc_results
        
        # Verdict Banner (Combined)
        target_f_name = f"{int(target_f)}F"
        req_prd = None
        if res is not None and not res[res['층이름'] == target_f_name].empty:
            req_prd = res[res['층이름'] == target_f_name].iloc[0]['필요_PRD_축력']
            
        banner_html = "<div class='ok-banner'>"
        
        if last_ok != "시공 불가":
            disp_name = last_ok.replace("F", "층")
            if disp_name.startswith("B"):
                disp_name = f"지하 {disp_name[1:]}"
            else:
                disp_name = f"지상 {disp_name}"
            banner_html += f"PRD <b>{limit:,.0f} kN</b>에서 기초 타설 전 <b style='color:red;'>{disp_name} 바닥</b>까지 시공 가능합니다."
        else:
            banner_html += f"PRD <b>{limit:,.0f} kN</b>에서 기초 타설 전 <b style='color:red;'>시공 불가</b>합니다."
            
        if req_prd is not None:
            if limit >= req_prd:
                banner_html += f"<div style='margin-top:8px; font-size:1.05rem; color:#334155;'> <b>지상 {int(target_f)}층(목표층)</b>까지 PRD 축력은 충분합니다. (필요 최소 {req_prd:,.0f} kN)</div>"
            else:
                banner_html += f"<div style='margin-top:8px; font-size:1.05rem; color:#334155;'> <b>지상 {int(target_f)}층(목표층)</b> 바닥까지 시공하려면 최소 <span style='color:#0284c7; font-weight:800;'>{req_prd:,.0f} kN</span> 의 PRD 축력이 필요합니다.</div>"
                
        banner_html += "</div>"
        st.markdown(banner_html, unsafe_allow_html=True)
        # Popover Navigation Bar
        p_col1, p_col2, p_col3 = st.columns([1.2, 1.4, 3])
        with p_col1:
            with st.popover("하중 비율", use_container_width=True):
                # Calculate total breakdown for floors marked as OK
                active_res = res[res["판정"] == "OK"]
                if not active_res.empty:
                    t_slab = active_res["슬래브 무게(kN)"].sum()
                    t_beam = active_res["보 무게(kN)"].sum()
                    t_col = active_res["기둥 무게(kN)"].sum()
                    t_add = active_res["추가하중(kN)"].sum()
                    t_ll = active_res["단일 LL(kN)"].sum()
                    t_total = t_slab + t_beam + t_col + t_add + t_ll
                    
                    if t_total > 0:
                        # Table display
                        ratio_df = pd.DataFrame({
                            "항목": ["슬래브(DL)", "보(DL)", "기둥(DL)", "기타 하중(DL)", "활하중(LL)"],
                            "하중 (kN)": [t_slab, t_beam, t_col, t_add, t_ll],
                            "비율 (%)": [
                                (t_slab/t_total)*100,
                                (t_beam/t_total)*100,
                                (t_col/t_total)*100,
                                (t_add/t_total)*100,
                                (t_ll/t_total)*100
                            ]
                        })
                        # Format and Show
                        st.markdown("<p style='font-weight:bold; margin-bottom:10px;'>하중 비율</p>", unsafe_allow_html=True)
                        st.table(ratio_df.style.format({"하중 (kN)": "{:,.1f}", "비율 (%)": "{:.1f}%"}))
                        st.markdown(f"<p style='text-align:right; font-size:0.9rem; color:#475569;'><b>합계: {t_total:,.1f} kN</b></p>", unsafe_allow_html=True)
                    else:
                        st.warning("하중 계산값이 0입니다.")
                else:
                    st.info("시공 가능한 층이 없어 비율을 표시할 수 없습니다.")
        
        with p_col2:
            with st.popover("시공 가능 층수 확인", use_container_width=True):
                st.pyplot(draw_professional_diagram(res, limit))
        
        st.write("")
        
        st.write("")
        
        st.write("")

        # Editor Section First
        st.markdown("<div class='section-title'>층별 입력 내역</div>", unsafe_allow_html=True)
        tab_sheet, tab_override, tab_details = st.tabs(["층별 내역", "특정 층 수정", "상세 계산 근거"])
        
        with tab_sheet:
            edit_df = res.copy()
            
            # Mapping internal to requested names
            column_mapping = {
                "층이름": "층",
                "층고(m)": "층고\n(m)",
                "X분담(m)": "X분담\n(m)",
                "Y분담(m)": "Y분담\n(m)",
                "면적": "면적\n(m2)",
                "Slab(m)": "슬래브\n(m)",
                "슬래브 무게(kN)": "슬래브\n무게(kN)",
                "보 무게(kN)": "보 무게\n(kN)",
                "기둥 무게(kN)": "기둥 무게\n(kN)",
                "추가하중(kN/m2)": "기타\n(kN/m2)",
                "LL(kN/m2)": "LL\n(kN/m2)",
                "단일 DL(kN)": "DL\n(kN)",
                "단일 LL(kN)": "LL\n(kN)",
                "단일 DL+LL(kN)": "🟡 DL+LL\n(kN)",
                "단일 1.2D+1.6L(kN)": "🟡 1.2D+1.6L\n(kN)",
                "단일 1.2D+1.28L(kN)": "🟡 1.2D+1.28L\n(kN)",
                "DL+LL": "🟠 누적\nDL+LL(kN)",
                "1.2D + 1.6L": "🟠 누적\n1.2D+1.6L",
                "1.2D + 1.28L": "🟠 누적\n1.2D+1.28L"
            }
            # Prepare DataFrame for Editor
            display_df = edit_df[list(column_mapping.keys())].rename(columns=column_mapping)
            
            # Add Total Sum Row
            sum_cols = ["슬래브\n무게(kN)", "보 무게\n(kN)", "기둥 무게\n(kN)", "DL\n(kN)", "LL\n(kN)", "🟡 DL+LL\n(kN)", "🟡 1.2D+1.6L\n(kN)", "🟡 1.2D+1.28L\n(kN)"]
            cum_cols = ["🟠 누적\nDL+LL(kN)", "🟠 누적\n1.2D+1.6L", "🟠 누적\n1.2D+1.28L"]
            total_sum_val = {col: display_df[col].sum() if col in sum_cols else None for col in display_df.columns}
            
            # For cumulative columns, show the final value (bottom-most floor)
            for c_col in cum_cols:
                if c_col in display_df.columns:
                    valid_vals = display_df[c_col].dropna()
                    if not valid_vals.empty:
                        total_sum_val[c_col] = valid_vals.iloc[-1]

            total_sum_val["층"] = "📊 합계"
            display_df = pd.concat([display_df, pd.DataFrame([total_sum_val])], ignore_index=True)
            
            # Configure Layout (Simplified Schema)
            res_cols = [
                "면적\n(m2)", "슬래브\n무게(kN)", "보 무게\n(kN)", "기둥 무게\n(kN)", 
                "DL\n(kN)", "LL\n(kN)", "🟡 DL+LL\n(kN)", "🟡 1.2D+1.6L\n(kN)", "🟡 1.2D+1.28L\n(kN)",
                "🟠 누적\nDL+LL(kN)", "🟠 누적\n1.2D+1.6L", "🟠 누적\n1.2D+1.28L"
            ]
            edited = st.data_editor(
                display_df, 
                hide_index=True, 
                use_container_width=True, 
                disabled=res_cols + ["층"],
                column_config={
                    "층": st.column_config.TextColumn(width=100),
                    "층고\n(m)": st.column_config.NumberColumn(width=80),
                    "X분담\n(m)": st.column_config.NumberColumn(width=80),
                    "Y분담\n(m)": st.column_config.NumberColumn(width=80),
                    "면적\n(m2)": st.column_config.NumberColumn(width=90),
                    "슬래브\n(m)": st.column_config.NumberColumn(width=80, format="%.2f"),
                    "슬래브\n무게(kN)": st.column_config.NumberColumn(width=100),
                    "보 무게\n(kN)": st.column_config.NumberColumn(width=100),
                    "기둥 무게\n(kN)": st.column_config.NumberColumn(width=100),
                    "기타\n(kN/m2)": st.column_config.NumberColumn(width=90),
                    "LL\n(kN/m2)": st.column_config.NumberColumn(width=90),
                    "DL\n(kN)": st.column_config.NumberColumn(width=110),
                    "LL\n(kN)": st.column_config.NumberColumn(width=110),
                    "🟡 DL+LL\n(kN)": st.column_config.NumberColumn(width=120),
                    "🟡 1.2D+1.6L\n(kN)": st.column_config.NumberColumn(width=130),
                    "🟡 1.2D+1.28L\n(kN)": st.column_config.NumberColumn(width=130),
                    "🟠 누적\nDL+LL(kN)": st.column_config.NumberColumn(width=130, format="%.1f"),
                    "🟠 누적\n1.2D+1.6L": st.column_config.NumberColumn(width=130, format="%.1f"),
                    "🟠 누적\n1.2D+1.28L": st.column_config.NumberColumn(width=130, format="%.1f"),
                }
            )
            
            if st.button("표 수정사항 적용"):
                # Identify edited rows, ignoring the 'Total Sum' row at the end
                n_floors = len(st.session_state.prd_df)
                floor_edited = edited.head(n_floors)
                floor_original = display_df.head(n_floors)
                
                diff = floor_edited.ne(floor_original).any(axis=1)
                changed_indices = diff[diff].index.tolist()
                
                if changed_indices:
                    for i in changed_indices:
                        if i < n_floors:
                            row = floor_edited.loc[i]
                            st.session_state.prd_df[i].update({
                                "층고(m)": float(row["층고\n(m)"]),
                                "X분담(m)": float(row["X분담\n(m)"]), 
                                "Y분담(m)": float(row["Y분담\n(m)"]), 
                                "LL(kN/m2)": float(row["LL\n(kN/m2)"]),
                                "Slab(m)": float(row["슬래브\n(m)"]),
                                "추가하중(kN/m2)": float(row["기타\n(kN/m2)"])
                            })
                            st.session_state.prd_df[i]["is_edited"] = True
                    
                    res, last_ok, logs = PRDEngine.run_analysis(st.session_state.prd_df, limit)
                    st.session_state.calc_results = (res, last_ok, logs)
                    st.rerun()

        with tab_override:
            oc1, oc2 = st.columns([1.5, 3.5])
            with oc1:
                target_f = st.selectbox("수정할 층 선택", [d["층이름"] for d in st.session_state.prd_df])
                
                # List of edited floors
                edited_floors = [d for d in st.session_state.prd_df if d.get("is_edited")]
                
                # Check if default sets are available for comparison. 
                # If they are None (e.g. data loaded from Excel), skip showing diffs against defaults.
                if edited_floors and f_set is not None and b_set is not None:
                    st.markdown("<p style='font-size:0.85rem; font-weight:bold; margin-top:20px; border-bottom:1px solid #e2e8f0; padding-bottom:5px;'>수정된 층 목록</p>", unsafe_allow_html=True)
                    for ef_data in edited_floors:
                        ef_name = ef_data["층이름"]
                        # Get default set for comparison
                        is_f = ef_name.endswith("F") and not ef_name.startswith("B")
                        s_def = f_set if is_f else b_set
                        
                        lc1, lc2 = st.columns([4, 1])
                        with lc1:
                            with st.expander(f"📍 {ef_name} 변경사항", expanded=False):
                                # Compare values and show only diffs
                                diff_found = False
                                for key_map, label in {
                                    "층고(m)": "층고", "X분담(m)": "X분담", "Y분담(m)": "Y분담",
                                    "LL(kN/m2)": "활하중", "Slab(m)": "슬래브", "추가하중(kN/m2)": "기타하중"
                                }.items():
                                    # Default value mapping
                                    def_val = s_def['h'] if key_map == "층고(m)" else (
                                              s_def['x'] if key_map == "X분담(m)" else (
                                              s_def['y'] if key_map == "Y분담(m)" else (
                                              s_def['ll'] if key_map == "LL(kN/m2)" else (
                                              s_def['s'] if key_map == "Slab(m)" else s_def['add_ll']))))
                                    
                                    # Special case for 1F LL
                                    if ef_name == "1F" and key_map == "LL(kN/m2)": def_val = 25.0
                                    
                                    if ef_data[key_map] != def_val:
                                        st.markdown(f"<span style='font-size:0.8rem; color:#64748b;'>{label}:</span> <span style='font-size:0.8rem; text-decoration:line-through; color:#94a3b8;'>{def_val}</span> → <span style='font-size:0.8rem; font-weight:bold; color:#0f172a;'>{ef_data[key_map]}</span>", unsafe_allow_html=True)
                                        diff_found = True
                                
                                # Column diff
                                if ef_data["c_type"] != s_def["ct"]:
                                    st.markdown(f"<span style='font-size:0.8rem; color:#64748b;'>기둥타입:</span> <span style='font-size:0.8rem; font-weight:bold; color:#0f172a;'>{ef_data['c_type']}</span>", unsafe_allow_html=True)
                                    diff_found = True
                                else:
                                    # Same type, check dimensions
                                    cv_def = s_def["cv"]
                                    cv_now = ef_data["c_vals"]
                                    c_diffs = []
                                    if ef_data["c_type"] == "RC":
                                        if cv_now.get('b') != cv_def.get('b'): c_diffs.append(f"B:{cv_def.get('b')}→{cv_now.get('b')}")
                                        if cv_now.get('h') != cv_def.get('h'): c_diffs.append(f"H:{cv_def.get('h')}→{cv_now.get('h')}")
                                    elif ef_data["c_type"] == "SRC":
                                        if cv_now.get('b') != cv_def.get('b'): c_diffs.append(f"B:{cv_def.get('b')}→{cv_now.get('b')}")
                                        if cv_now.get('h') != cv_def.get('h'): c_diffs.append(f"H:{cv_def.get('h')}→{cv_now.get('h')}")
                                        if cv_now.get('tw') != cv_def.get('tw'): c_diffs.append(f"tw:{cv_def.get('tw')}→{cv_now.get('tw')}")
                                    
                                    if c_diffs:
                                        st.markdown(f"<span style='font-size:0.8rem; color:#64748b;'>기둥규격:</span> <span style='font-size:0.8rem; font-weight:bold; color:#0f172a;'>{', '.join(c_diffs)}</span>", unsafe_allow_html=True)
                                        diff_found = True

                                # Beam diff
                                b_def = s_def["b_list"]
                                b_now = ef_data["b_list"]
                                if len(b_now) != len(b_def):
                                    st.markdown(f"<span style='font-size:0.8rem; color:#64748b;'>보 개수:</span> <span style='font-size:0.8rem; font-weight:bold; color:#0f172a;'>{len(b_def)}개 → {len(b_now)}개</span>", unsafe_allow_html=True)
                                    diff_found = True
                                else:
                                    # Compare items
                                    b_changed = False
                                    for i in range(len(b_now)):
                                        if b_now[i] != b_def[i]:
                                            b_changed = True; break
                                    if b_changed:
                                        st.markdown(f"<span style='font-size:0.8rem; color:#64748b;'>보 상세정보 수정됨</span>", unsafe_allow_html=True)
                                        diff_found = True
                                
                                if not diff_found:
                                    st.markdown("<span style='font-size:0.8rem; color:#94a3b8;'>표 데이터 외 미세 수정됨</span>", unsafe_allow_html=True)

                        if lc2.button("🗑️", key=f"del_{ef_name}", help=f"{ef_name} 초기화"):
                            # Reset to defaults
                            idx_reset = [i for i, d in enumerate(st.session_state.prd_df) if d["층이름"] == ef_name][0]
                            st.session_state.prd_df[idx_reset].update({
                                "층고(m)": s_def['h'], "X분담(m)": s_def['x'], "Y분담(m)": s_def['y'],
                                "LL(kN/m2)": 25.0 if ef_name == "1F" else s_def['ll'],
                                "Slab(m)": s_def['s'], "추가하중(kN/m2)": s_def['add_ll'],
                                "c_type": s_def['ct'], "c_vals": s_def['cv'].copy(), 
                                "b_list": [b.copy() for b in s_def['b_list']] if isinstance(s_def['b_list'], list) else [],
                                "is_edited": False
                            })
                            res, last_ok, logs = PRDEngine.run_analysis(st.session_state.prd_df, limit)
                            st.session_state.calc_results = (res, last_ok, logs); st.rerun()

            idx = [i for i, d in enumerate(st.session_state.prd_df) if d["층이름"] == target_f][0]
            ex = st.session_state.prd_df[idx]
            
            with oc2:
                m_set = member_input_form_compact(f"M_{target_f}", defaults={'h':ex['층고(m)'], 'x':ex['X분담(m)'], 'y':ex['Y분담(m)'], 's':ex['Slab(m)'], 'll':ex['LL(kN/m2)'], 'add_ll':ex['추가하중(kN/m2)'], 'ct':ex['c_type'], 'cv':ex['c_vals'], 'b_list':ex['b_list']})
                if st.button(f"{target_f} 업데이트 적용", key="ovr_btn"):
                    st.session_state.prd_df[idx].update({
                        "층고(m)":m_set['h'], "X분담(m)":m_set['x'], "Y분담(m)":m_set['y'], 
                        "Slab(m)":m_set['s'], "LL(kN/m2)":m_set['ll'], "추가하중(kN/m2)":m_set['add_ll'], 
                        "c_type":m_set['ct'], "c_vals":m_set['cv'], "b_list":m_set['b_list'], "is_edited": True
                    })
                    res, last_ok, logs = PRDEngine.run_analysis(st.session_state.prd_df, limit)
                    st.session_state.calc_results = (res, last_ok, logs); st.rerun()

        with tab_details:
             for n, l in logs.items():
                with st.expander(f"📄 {n} 상세 계산 근거"):
                    c_a, c_b = st.columns([1.5, 2.5])
                    with c_a:
                        dl = l["DL_Sum"]
                        ll = l["LL_Sum"]
                        st.markdown(f"** DL:** {dl:,.2f} kN")
                        st.markdown(f"** LL:** {ll:,.2f} kN")
                        st.markdown("<hr style='margin:10px 0; border-top:1px solid #e2e8f0;'>", unsafe_allow_html=True)
                        st.markdown(f"**🟡 DL+LL:** {dl+ll:,.2f} kN")
                        st.markdown(f"**🟡 1.2D + 1.6L:** {1.2*dl + 1.6*ll:,.2f} kN")
                        st.markdown(f"**🟡 1.2D + 1.28L:** {1.2*dl + 1.28*ll:,.2f} kN")
                    with c_b:
                        for log in l["Logs"]: st.code(log, language="markdown")



        st.markdown("---")
        
        # 1. Export to Excel
        if st.session_state.prd_df and res is not None:
            # Prepare rich data for export (inputs + calculation results)
            export_raw = res.copy()
            
            # Add flattened member info to the result dataframe
            for i, row in export_raw.iterrows():
                cv = row['c_vals']
                if isinstance(cv, dict):
                    for k, v in cv.items():
                        export_raw.loc[i, f"col_{k}"] = v
                export_raw.loc[i, "beam_data_json"] = json.dumps(row['b_list'])
            
            # Define Column Labels (Inputs + Calculations)
            final_excel_cols = {
                "층이름": "층",
                "층고(m)": "층고(m) [입력]",
                "X분담(m)": "X분담(m) [입력]",
                "Y분담(m)": "Y분담(m) [입력]",
                "Slab(m)": "슬래브두께(m) [입력]",
                "LL(kN/m2)": "입력LL(kN/m2) [입력]",
                "추가하중(kN/m2)": "추가DL(kN/m2) [입력]",
                "c_type": "기둥유형 [입력]",
                "단일 DL(kN)": "단일 DL(kN) [결과]",
                "단일 LL(kN)": "단일 LL(kN) [수식]",
                "단일 DL+LL(kN)": "단일 DL+LL(kN) [수식]",
                "DL+LL": "누적 DL+LL(kN) [수식]",
                "1.2D + 1.6L": "누적 1.2D+1.6L(kN) [수식]",
                "1.2D + 1.28L": "누적 1.2D+1.28L(kN) [수식]",
                "판정": "판정_Helper" # Hidden helper for logic
            }

            # Columns that are editable by user
            editable_cols = [
                "층고(m) [입력]", "X분담(m) [입력]", "Y분담(m) [입력]", 
                "슬래브두께(m) [입력]", "입력LL(kN/m2) [입력]", "추가DL(kN/m2) [입력]", "기둥유형 [입력]"
            ]
            
            tech_cols = [col for col in export_raw.columns if col.startswith("col_") or col == "beam_data_json"]
            cols_to_keep = [c for c in final_excel_cols.keys() if c in export_raw.columns]
            excel_display = export_raw[cols_to_keep + tech_cols].rename(columns=final_excel_cols)

            # Prepare column lookup
            col_names_list = excel_display.columns.tolist()
            def gcl(idx): return chr(65 + idx) if idx < 26 else (chr(65 + (idx//26)-1) + chr(65 + (idx%26)))
            
            def get_col_idx(name):
                try: return col_names_list.index(name)
                except ValueError: return -1

            # Key Columns
            ix_x_idx = get_col_idx("X분담(m) [입력]")
            ix_y_idx = get_col_idx("Y분담(m) [입력]")
            ix_ll_u_idx = get_col_idx("입력LL(kN/m2) [입력]")
            ix_dl_s_idx = get_col_idx("단일 DL(kN) [결과]")
            ix_ll_s_idx = get_col_idx("단일 LL(kN) [수식]")
            ix_sum_s_idx = get_col_idx("단일 DL+LL(kN) [수식]")
            ix_verdict_idx = get_col_idx("판정_Helper")
            
            # Convert to letters for Excel formulas (if found)
            ix_x = gcl(ix_x_idx) if ix_x_idx != -1 else "A"
            ix_y = gcl(ix_y_idx) if ix_y_idx != -1 else "A"
            ix_ll_u = gcl(ix_ll_u_idx) if ix_ll_u_idx != -1 else "A"
            ix_dl_s = gcl(ix_dl_s_idx) if ix_dl_s_idx != -1 else "A"
            ix_ll_s = gcl(ix_ll_s_idx) if ix_ll_s_idx != -1 else "A"
            ix_sum_s = gcl(ix_sum_s_idx) if ix_sum_s_idx != -1 else "A"
            ix_verdict = gcl(ix_verdict_idx) if ix_verdict_idx != -1 else "A"
            
            # Calculation Row Mapping (Data starts at row 4 due to PRD Header)
            data_start_row = 4
            n_rows = len(excel_display)
            limit_cell = "$C$1"
            
            # Insert Formulas for Dynamic PRD Linkage
            for i in range(n_rows):
                r = i + data_start_row
                # 1. Single LL
                excel_display.iloc[i, col_names_list.index("단일 LL(kN) [수식]")] = f"={ix_x}{r}*{ix_y}{r}*{ix_ll_u}{r}"
                # 2. Single DL+LL
                excel_display.iloc[i, col_names_list.index("단일 DL+LL(kN) [수식]")] = f"={ix_dl_s}{r}+{ix_ll_s}{r}"
                
                # 3. Verdict Helper (Bottom-up Logic)
                # Floor is OK if (Load of this floor + loads below <= Limit) AND (Floor below is OK)
                # Foundation (last row) is always contextually OK for the iteration to start
                if i == n_rows - 1: # Last row (Foundation)
                    # Foundation usually doesn't count towards PRD limit in this specific engine, but let's assume always contextually OK
                    excel_display.iloc[i, col_names_list.index("판정_Helper")] = 1
                elif i == n_rows - 2: # Bottom-most floor to check
                    excel_display.iloc[i, col_names_list.index("판정_Helper")] = f"=IF({ix_sum_s}{r}<={limit_cell}, 1, 0)"
                else:
                    # Current floor OK if Cumulative sum below <= limit AND floor directly below is OK
                    excel_display.iloc[i, col_names_list.index("판정_Helper")] = f"=IF(AND(SUM({ix_sum_s}{r}:{ix_sum_s}{n_rows+data_start_row-2})<={limit_cell}, {ix_verdict}{r+1}=1), 1, 0)"

                # 4. Cumulative Sums (Simplified using SUMIFS for maximum compatibility)
                # Since NG is always at the top, SUMIFS(range, verdict_range, 1) correctly starts from the first OK row.
                ix_idx_sum_cum = get_col_idx("누적 DL+LL(kN) [수식]")
                ix_idx_12d16l_cum = get_col_idx("누적 1.2D+1.6L(kN) [수식]")
                ix_idx_12d128l_cum = get_col_idx("누적 1.2D+1.28L(kN) [수식]")
                
                # Definitions for SUMIFS
                v_fixed_range = f"${ix_verdict}${data_start_row}:{ix_verdict}{r}"
                sum_s_fixed_range = f"${ix_sum_s}${data_start_row}:{ix_sum_s}{r}"
                dl_s_fixed_range = f"${ix_dl_s}${data_start_row}:{ix_dl_s}{r}"
                ll_s_fixed_range = f"${ix_ll_s}${data_start_row}:{ix_ll_s}{r}"
                
                # Common condition: verdict must be 1 (OK)
                if ix_idx_sum_cum != -1:
                    excel_display.iloc[i, ix_idx_sum_cum] = f"=IF({ix_verdict}{r}=1, SUMIFS({sum_s_fixed_range}, {v_fixed_range}, 1), 0)"
                if ix_idx_12d16l_cum != -1:
                    excel_display.iloc[i, ix_idx_12d16l_cum] = f"=IF({ix_verdict}{r}=1, 1.2*SUMIFS({dl_s_fixed_range}, {v_fixed_range}, 1) + 1.6*SUMIFS({ll_s_fixed_range}, {v_fixed_range}, 1), 0)"
                if ix_idx_12d128l_cum != -1:
                    excel_display.iloc[i, ix_idx_12d128l_cum] = f"=IF({ix_verdict}{r}=1, 1.2*SUMIFS({dl_s_fixed_range}, {v_fixed_range}, 1) + 1.28*SUMIFS({ll_s_fixed_range}, {v_fixed_range}, 1), 0)"

            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # First, write at Row 4
                excel_display.to_excel(writer, index=False, sheet_name='PRD_Analysis', startrow=2)
                
                workbook = writer.book
                worksheet = writer.sheets['PRD_Analysis']
                
                # 5. HEADER: PRD Setup (Row 1)
                worksheet.merge_cells('A1:B1')
                worksheet['A1'] = "PRD 허용축력 ="
                worksheet['A1'].font = Font(bold=True)
                worksheet['C1'] = limit # Current Limit
                worksheet['C1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow
                worksheet['D1'] = "kN"
                worksheet['A1'].alignment = Alignment(horizontal='right')
                
                # 6. TOTAL SUM ROW
                last_data_row = n_rows + data_start_row - 1
                sum_row_num = last_data_row + 1
                worksheet[f'A{sum_row_num}'] = "📊 합계"
                worksheet[f'A{sum_row_num}'].font = Font(bold=True)
                
                # Sum columns (DL, LL Single, etc.)
                for col_idx in range(len(col_names_list)):
                    col_letter = gcl(col_idx)
                    col_name = col_names_list[col_idx]
                    if "[결과]" in col_name or "[수식]" in col_name:
                        # For cumulative cols, show the last value
                        if "누적" in col_name:
                            worksheet[f'{col_letter}{sum_row_num}'] = f"={col_letter}{last_data_row}"
                        else:
                            worksheet[f'{col_letter}{sum_row_num}'] = f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})"
                        worksheet[f'{col_letter}{sum_row_num}'].font = Font(bold=True)
                
                # 7. LEGEND / NOTE (Bottom)
                note_row = sum_row_num + 2
                worksheet.merge_cells(f'A{note_row}:F{note_row}')
                worksheet[f'A{note_row}'] = "⚠️ 사용 안내 및 범례"
                worksheet[f'A{note_row}'].font = Font(bold=True, size=11)
                
                worksheet[f'A{note_row+1}'] = "■ 흰색 셀: 사용자 입력이 가능한 구간입니다. (층고, 면적, 하중 등)"
                worksheet[f'A{note_row+2}'] = "■ 짙은 회색 셀: 시스템 계산 구간입니다. 수동 수정을 권장하지 않습니다."
                worksheet[f'A{note_row+3}'] = f"■ 노란색 셀(C1): PRD 허용축력을 변경하면 시공 가능 범위와 누적 하중이 자동 갱신됩니다."
                worksheet[f'A{note_row+4}'] = "■ 기술 열(Hidden): 'beam_data_json' 등 숨겨진 열을 삭제하면 프로그램 불러오기가 불가능합니다."
                
                # Styles
                header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                input_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") 
                result_fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid") 
                white_font = Font(color="FFFFFF")
                black_font = Font(color="000000")
                
                # Format Table Header (Row 3)
                for cell in worksheet[3]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Apply Styles to Body
                for r_idx in range(data_start_row, sum_row_num + 1):
                    for c_idx, col_name in enumerate(col_names_list):
                        cell = worksheet.cell(row=r_idx, column=c_idx+1)
                        if col_name in editable_cols:
                            cell.fill = input_fill
                            cell.font = black_font
                        elif "[결과]" in col_name or "[수식]" in col_name:
                            cell.fill = result_fill
                            cell.font = white_font
                        elif col_name == "층":
                            cell.font = Font(bold=True)
                        
                        cell.border = Border(left=Side(style='thin', color="CBD5E1"),
                                           right=Side(style='thin', color="CBD5E1"),
                                           top=Side(style='thin', color="CBD5E1"),
                                           bottom=Side(style='thin', color="CBD5E1"))

                # Hide Tech & Helper Columns
                for c_idx, col_name in enumerate(col_names_list):
                    if col_name in tech_cols or col_name == "판정_Helper":
                        worksheet.column_dimensions[gcl(c_idx)].visible = False
                
                # Auto-adjust column width (Robust against merged cells)
                for i_col, col in enumerate(worksheet.columns, 1):
                    max_length = 0
                    column_letter = get_column_letter(i_col)
                    for cell in col:
                        # Skip if it's a MergedCell or has no value property accessibly
                        try:
                            if cell.value:
                                val_len = len(str(cell.value))
                                if val_len > max_length:
                                    max_length = val_len
                        except: pass
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 35)

            st.download_button(
                label="엑셀 저장",
                data=buffer.getvalue(),
                file_name="PRD_Report_Interactive.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.button("엑셀 저장", disabled=True)
    else:
        # WELCOME SCREEN
        st.markdown(f"""
            <div style='text-align:center; padding: 40px 20px; color: #1e293b; background: white; border-radius: 20px; border: 2px dashed #e2e8f0;'>
                <h2 style='font-weight: 800; margin-bottom: 15px;'>PRD 축력 검토</h2>
                <p style='color: #64748b; font-size: 1.05rem; max-width: 600px; margin: 0 auto;'>
                    왼쪽 사이드바에서 프로젝트 층수와 허용 축력을 입력한 후 <b>해석 실행 (RUN ANALYSIS)</b> 버튼을 눌러
                    정밀한 구조 검토 레포트를 생성하세요.
                </p>
            </div>
            """, unsafe_allow_html=True)
        
        if 'show_guide' not in st.session_state:
            st.session_state.show_guide = False
            
        st.write("")
        c1, c2, c3 = st.columns([1, 1.2, 1])
        with c2:
            if st.button("역타 공법 개념도 보기 / 닫기", use_container_width=True):
                st.session_state.show_guide = not st.session_state.show_guide

        if st.session_state.show_guide:
            st.markdown("<div style='text-align:center; margin-top: 20px;'><h4 style='color: #475569;'>역타 공법 개념도</h4></div>", unsafe_allow_html=True)
            gc1, gc2 = st.columns(2)
            with gc1:
                st.image(os.path.join(os.path.dirname(__file__), "슬라이드1.JPG"), caption="[입력값] 산출 위치 가이드", use_container_width=True)
            with gc2:
                st.image(os.path.join(os.path.dirname(__file__), "슬라이드2.JPG"), caption="[검토값] 산출 위치 가이드", use_container_width=True)

    # Core logic trigger
    if st.session_state.get('run_analysis'):
        names = ["Roof"] + [f"{i}F" for i in range(int(f_cnt), 0, -1)] + [f"B{i}F" for i in range(1, int(b_cnt) + 1)]
        current_data = {d['층이름']: d for d in st.session_state.prd_df} if st.session_state.prd_df else {}
        new_data = []
        for n in names:
            if n in current_data and current_data[n].get('is_edited'):
                new_data.append(current_data[n])
            else:
                is_f = n.endswith("F") and not n.startswith("B")
                s = f_set if is_f else b_set
                
                # 1F Default Load Logic (25.0)
                if n == "1F":
                    initial_ll = 25.0
                else:
                    initial_ll = s['ll']
                
                # Ensure deep copies of nested structures, handle potential string leftovers
                cv_data = s['cv'] if isinstance(s['cv'], dict) else {}
                bl_data = s['b_list'] if isinstance(s['b_list'], list) else []
                
                new_data.append({
                    "층이름": n, "층고(m)": s['h'], "X분담(m)": s['x'], "Y분담(m)": s['y'],
                    "LL(kN/m2)": initial_ll, "Slab(m)": s['s'], "추가하중(kN/m2)": s['add_ll'],
                    "c_type": s['ct'], "c_vals": cv_data.copy(), "b_list": [b.copy() for b in bl_data], "is_edited": False
                })
        
        st.session_state.prd_df = new_data
        res, last_ok, logs = PRDEngine.run_analysis(new_data, limit)
        st.session_state.calc_results = (res, last_ok, logs)
        st.session_state['run_analysis'] = False
        st.rerun()

if __name__ == "__main__":
    main()